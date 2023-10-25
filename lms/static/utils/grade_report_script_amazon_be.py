# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import zipfile

import sys
importlib.reload(sys)

import os
from io import BytesIO

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################



import time
import datetime
import json

from opaque_keys.edx.locations import SlashSeparatedCourseKey
from opaque_keys.edx import locator
from lms.djangoapps.wul_apps.models import WulCourseEnrollment
from opaque_keys.edx.locator import CourseLocator
from lms.djangoapps.courseware.courses import get_course_by_id

from lms.djangoapps.wul_apps.best_grade.helpers import check_best_grade

from common.djangoapps.student.models import User, UserProfile
from lms.djangoapps.courseware.models import StudentModule
from student.models import CourseEnrollment


from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import logging
log = logging.getLogger()



emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")

## Workbook
wb = Workbook()
ws = wb.active
ws.title = "Grade report"


all_courses_video_student_module = []

for course_id in course_ids :
    list_of_student_modules = StudentModule.objects.filter(course_id__exact=course_id, module_type__exact="video").order_by().values('student_id', 'module_state_key', 'state')

    for student_module in list_of_student_modules :
        all_courses_video_student_module.append(student_module)


# Must add the old course data : course-v1:amazon+amazon001+SP

list_of_old_videos = StudentModule.objects.filter(course_id__exact="course-v1:amazon+amazon001+SP", module_type__exact="video").order_by().values('student_id', 'module_state_key', 'state')
for student_module in list_of_old_videos :
    all_courses_video_student_module.append(student_module)



# Set correspondance table
with open('/home/ubuntu/rapports_amazon_table_be.json') as json_file:
    idVideoCorrespondance = json.load(json_file)

videos_list = list(idVideoCorrespondance.keys())

## Construct data
users = User.objects.all()
users_data = dict()
siret = dict()

# Headers
headers = ["Username","Email","Nom complet","Region","Siret","Numéro de téléphone","Vendez-vous en ligne","Date de création de compte", "Date de dernière connexion","Temps passé - total","Nombre de cours suivis","Nombre de cours validés","Introduction","Démarrer son activité","Préparer votre transition numérique","Vendre sur son site personnel","Vendre sur un site tiers", "Introductie", "Uw bedrijf starten", "Uw digitale overgang voorbereiden","Verkopen op je eigen website", "Verkopen op een site van derden"]

headers.append('Nombre total de vidéo vues')

for name in videos_list:
    # video_id = name.split("+")[4]
    video_id = name
    if video_id in idVideoCorrespondance :
        dict_content = idVideoCorrespondance.get(video_id, {})
        str_video_chapter = str(dict_content.get('numeroDeChapitre', 'n/a'))
        str_video_name = dict_content.get('nomDeVideo', 'n/a')
        str_video_course_number = str(dict_content.get('numeroDeCours', 'n/a'))
        str_video_course = dict_content.get('nomDuCours', 'n/a')
        headers.append("Cours  "+str_video_course_number+" - \""+str_video_course+"\", Chapitre "+str_video_chapter+" : \""+str_video_name+"\"")
    else :
        headers.append(name.split("+")[4])



today = datetime.datetime.now(datetime.timezone.utc)

def convert_str_to_obj(saved_video_position):
    str_time_video = saved_video_position.replace('"saved_video_position": "', '').replace('{', '').replace('}', '').replace('"', '')
    if 'speed' in str_time_video:
        char_to_replace = ["0.25", "0.5", "0.75","1.0","1.25","1.5","1.75","2.0", ",", " ","speed:"]
        for char in char_to_replace:
            str_time_video = str_time_video.replace(char, "")
    return str_time_video

def convert_str_to_int(time_str):
    """Get seconds from time."""
    h, m, s = time_str.split(':')
    return int(h) * 3600 + int(m) * 60 + int(s)



### Loop over all_user 


for index, user in enumerate(users):

    # uncomment this lines for testing, 
    # if index == 250:
    #     break


    # Escape fake email address
    if user.email.find("@example")!= -1 or user.email.find("@themoocagency") != -1 or user.email.find("@weuplearning")!= -1 or user.email.find("@yopmail")!= -1 or user.email.find("@amazon")!= -1 or user.email.find("@fake")!= -1:
        continue


    log.info('treating user :')
    log.info(user.email)


    user_data = dict()

    user_data["name"] = user.profile.name
    try:
        user_data["id"] = user.id
    except:
        user_data["id"] = ""
    try:
        user_data["username"] = user.username
    except:
        user_data["username"] = ""
    try:
        user_data["email"] = user.email
    except:
        user_data["email"] = ""
    custom_field = json.loads(user.profile.custom_field)
    try:
        user_data["region"] = custom_field["region"]
    except:
        user_data["region"] = ""
    try:
        user_data["siret"] = custom_field["company"]
    except:
        user_data["siret"] = ""
    try:
        user_data["phone_number"] = custom_field["phone_number"]
    except:
        user_data["phone_number"] = ""
    try:
        user_data["online_sales"] = custom_field["online_sales"]
    except:
        user_data["online_sales"] = ""
    try:
        user_data["date_joined"] = user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
    except:
        user_data["date_joined"] = ""
    try:
        user_data["last_login"] = user.last_login.strftime('%Y-%m-%d %H:%M:%S')
    except:
        user_data["last_login"] = ""


    user_row = []
    video_dict = dict()
    user_data["enrolled_to"] = 0
    user_data["finished_course"] = 0
    user_data["total_video_views"] = 0


    ### Grade Data

    try:
        global_time_tracking_cumul = custom_field["total_time_calculated"]
    except:
        global_time_tracking_cumul = 0


    for course_id in course_ids :

        all_course_enrollment = CourseEnrollment.objects.filter(user=user)
        user_data[course_id] = ''


        for enrollment in all_course_enrollment :

            log.info(str(enrollment.course_id))
            # if str(enrollment.course_id) == "course-v1:amazon+amazon001+SP" :
            #     continue


            if str(course_id) == str(enrollment.course_id) :

                course_key = CourseLocator.from_string(course_id)
                course = get_course_by_id(course_key)

                user_data["enrolled_to"] += 1

                log.info(course_id)
                try:
                    gradesTest = check_best_grade(user, course, force_best_grade=True)
                    user_data[course_id] = gradesTest.summary['percent']

                    if gradesTest.summary['percent'] >= 0.7 :
                        user_data["finished_course"] += 1
                except:
                    user_data[course_id] = 'Pas noté'


                try:
                    course_key = locator.CourseLocator.from_string(str(course_id))
                    wul_course_enrollment = WulCourseEnrollment.objects.get(course_enrollment_edx__user=user, course_enrollment_edx__course_id=course_key)

                    global_time_tracking = wul_course_enrollment.global_time_tracking
                    global_time_tracking_cumul += global_time_tracking
                except:
                    pass


    ### TimeTracking Data

    if global_time_tracking_cumul == 0 :
        user_data["global_time_tracking"] = 'n/a'
    else:
        user_data["global_time_tracking"] = datetime.timedelta(seconds=global_time_tracking_cumul)


    ### Video Data 


    for video in videos_list :
        video_dict[video] = "Non"

    # total_video_seconds = 0
    try:


        for result in all_courses_video_student_module:
            if(user.id == result["student_id"]):
                user_data["total_video_views"] += 1

                if str(result["module_state_key"]).split("+")[4] in video_dict:
                    user_state = result["state"]

                    video_time_tracking = convert_str_to_obj(user_state)

                    time_video_time_tracking = convert_str_to_int(video_time_tracking)
                    # total_video_seconds += time_video_time_tracking

                    video_dict[str(result["module_state_key"]).split("+")[4]] = datetime.timedelta(seconds=time_video_time_tracking)

                    # Optimize the loop
                    all_courses_video_student_module.remove(result)
                else:
                    print('ERROR {} VIDEO NOT IN THE LIST !'.formet(video_dict[str(result["module_state_key"]).split("+")[4]]))


    except:
        print(str(user.id)+" : error with watch a video field")
        # user_data["watch_a_video"] = ""



    user_row = [user_data["username"],user_data["email"],user_data["name"],user_data["region"],user_data["siret"],user_data["phone_number"],user_data['online_sales'],user_data["date_joined"],user_data["last_login"],user_data["global_time_tracking"],user_data["enrolled_to"],user_data["finished_course"]]

    for course_id in course_ids :
        user_row.append(user_data[course_id])


    user_row.append(user_data["total_video_views"])
    for video in video_dict :
        user_row.append(video_dict[video])
    # user_row.append(user_data["global_time_tracking"])



    users_data[user.username.capitalize()] = user_row

ordered_users = sorted(users_data.items(), key=lambda x: x[1])


### Print excel file


row = 1

sheet = wb.active
for i, header in enumerate(headers):
    sheet.cell(1, (i+1), header)
    sheet.cell(1, i+1).fill = PatternFill("solid", fgColor="1E2631")
    sheet.cell(1, i+1).font = Font(b=True, color="BA4926")
j=2

for user in ordered_users:
    user_row = user[1]
    l=0
    for value in user_row :
        sheet.cell(row=j, column=(l+1)).value = value
        l=l+1
    j=j+1

timestr = time.strftime("%Y_%m_%d")
filename = "Rapport_de_notes_Amazon_{}.xlsx".format(timestr)
filepath = '/home/ubuntu/amazon_reports/{}'.format(filename)
wb.save(filepath)


### Create a new zip file and write the Excel file into it


zipname = "rapport_de_notes.zip"
zippath = '/home/ubuntu/amazon_reports/{}'.format(zipname)

with zipfile.ZipFile(zippath, 'w', zipfile.ZIP_DEFLATED, compresslevel=9) as myzip:
    myzip.write(filepath, arcname=filename)

output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de donn&eacute;es Amazon<br/><br/>Bonne réception<br/>L'équipe WeupLearning"


### Send email


for email in emails:

    part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "Amazon <ne-pas-repondre@themoocagency.com>"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = email
    msg['Subject'] = "Rapport de notes Amazon"

    attachment = _files_values

    # Load your zip file instead of the Excel file
    with open(zippath, 'rb') as f:
        attachment = f.read()
        
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= {}".format(zipname))
    msg.attach(part)

    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, email, text)
    server.quit()

    print('Email sent to ',email)


## delete old files
two_weeks_ago = datetime.datetime.today() - datetime.timedelta(days=14)
try:
    os.remove('/home/ubuntu/amazon_reports/Rapport_de_notes_Amazon_{}.xlsx'.format(two_weeks_ago.strftime("%Y_%m_%d")))
except:
    pass




# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/amazon-belgique/lms/static/utils/grade_report_script_amazon_be.py "cyril.adolf@weuplearning.com" "course-v1:amazon+Introduction+AZ_01;course-v1:amazon+demarrer_activite+AZ_02;course-v1:amazon+Transition_numerique+AZ_03;course-v1:amazon+vendre_site_personnel+AZ_04;course-v1:amazon+vendre_site_tiers+AZ_05;course-v1:amazon+Introduction+AZNL_01;course-v1:amazon+demarrer_activite+AZNL_02;course-v1:amazon+Transition_numerique+AZNL_03;course-v1:amazon+vendre_site_personnel+AZNL_04;course-v1:amazon+vendre_site_tiers+AZNL_05"
