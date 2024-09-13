# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################


import json
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


# from datetime import datetime, date, timedelta
# from django.utils import timezone
# from dateutil import tz


from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from lms.djangoapps.courseware.courses import get_course_by_id
from lms.djangoapps.wul_apps.best_grade.helpers import check_best_grade

from openedx.core.djangoapps.site_configuration import helpers as configuration_helpers


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


import logging
log = logging.getLogger()


org = "sncf-voyageurs"
register_form = configuration_helpers.get_value_for_org(org, 'FORM_EXTRA')

# Get headers
HEADERS_GLOBAL = []
HEADERS_USER = [u"Email", u"Nom complet"]

HEADERS_FORM = []
NICE_HEADERS_FORM = []
if register_form is not None:
  for row in register_form:
    if row.get('type') is not None:
      HEADERS_FORM.append(row.get('name'))
      NICE_HEADERS_FORM.append(row.get('label'))


NICE_HEADER = list(NICE_HEADERS_FORM)
TECHNICAL_HEADER = list(HEADERS_FORM)



HEADERS_USER.extend(NICE_HEADER)
HEADERS_USER.append('Note obtenue (en %)')
HEADER = HEADERS_USER

emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")

all_users_data = {}


for course_id in course_ids:
  course_key = CourseLocator.from_string(course_id)
  course = get_course_by_id(course_key)
  course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)
  course_name = course.display_name_with_default

  course_data = {}

  for i in range(len(course_enrollments)):
    user = course_enrollments[i].user
    user_data = []

    enrollment = course_enrollments[i]

    user_CF_data = json.loads(user.profile.custom_field)

    if str(user.email).find('@yopmail') != -1 or str(user.email).find('@weuplearning') != -1 or str(user.email).find('@themoocagency') != -1 :
      continue


    try:
      user_data.append(user.email)
    except:
      try:
        user_data.append(user_CF_data['email'])
      except:
        user_data.append('n.a.')
    
    user_data.append(user.profile.name)



    for key in TECHNICAL_HEADER :

      try :
        user_data.append(user_CF_data[key])
      except :
        user_data.append('n.a.')


    # Grade
    gradesTest = check_best_grade(user, course, force_best_grade=True)
    userPersentGrade = gradesTest.summary['percent']*100

    try:
      user_data.append(userPersentGrade)
    except:
      user_data.append(0)

    data = { "general": user_data }
    course_data[str(user.id)] = data

  all_users_data[course_id]= course_data


# WRITE XLS
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= 'Rapport de notes'
filename = '/home/edxtma/csv/{}_sncf_voyageurs_grade_report.xlsx'.format(timestr)


correspondance_CF = {
  'true' : 'Oui',
  'false' : 'Non',
  'bac_5' : 'Bac +5',
  'bac_2_3' : 'Bac +2/+3',
  'bac' : 'Bac',
  'bac_general' : 'Bac général',
  'bac_technologique' : 'Bac technologique',
  'bacpro' : 'Bac professionnel',
  'bep_cap' : 'BEP/CAP',
  'no_level' : 'Pas de diplôme',
  'oui' : 'Oui',
  'non' : 'Non'
}


for i, header in enumerate(HEADER):
  sheet.cell(1, i+1, header)
  sheet.cell(1, i+1).font = Font(b=True, color="000000")


j=2

for k, course_id in all_users_data.items():
  
  for key, user in course_id.items():
    for i in range(len(user['general'])):

      if user['general'][i] in correspondance_CF:
        sheet.cell(j, i+1, correspondance_CF[user['general'][i]])
      else :
        sheet.cell(j, i+1, user['general'][i])

    j += 1


# SEND MAILS
# course_names = []
course_names_html = []
for course_id in course_ids: 
  course = get_course_by_id(CourseLocator.from_string(course_id)) 
  # course_names.append(course.display_name_with_default)
  course_names_html.append("<li>"+ str(course.display_name_with_default)+"</li>")

output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note concernant le cours : "+ course_names_html +"<br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

for email in emails:

  part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
  fromaddr = "ne-pas-repondre@themoocagency.com"
  msg = MIMEMultipart()
  msg['From'] = fromaddr
  msg['To'] = email
  msg['Subject'] = u"Rapport de notes Sureté SNCF"
  attachment = _files_values
  part = MIMEBase('application', 'octet-stream')
  part.set_payload(attachment)
  encoders.encode_base64(part)
  part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
  msg.attach(part)
  server = smtplib.SMTP('mail3.themoocagency.com', 25)
  server.starttls()
  server.login('contact', 'waSwv6Eqer89')
  msg.attach(part2)
  text = msg.as_string()
  server.sendmail(fromaddr, email, text)
  server.quit()
  log.info('Email sent to '+str(email))


log.info('------------> Finish calculate grades and write xlsx report')


# Qualif
# pas de cours .... 

# PROD
# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/sncf-voyageurs/lms/static/utils/grade_report_script.py 'cyril.adolf@weuplearning.com' course-v1:sncf-voyageurs+DC2+2024

