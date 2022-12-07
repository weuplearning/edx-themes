# -*- coding: utf-8 -*-
#!/usr/bin/env python
from distutils import core
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################


import json
import time
from openpyxl import Workbook
from datetime import datetime, date, timedelta
from django.utils import timezone
from dateutil import tz


from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from lms.djangoapps.courseware.courses import get_course_by_id
from lms.djangoapps.grades.context import grading_context_for_course
from lms.djangoapps.courseware.user_state_client import DjangoXBlockUserStateClient


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


import logging
log = logging.getLogger()

# 'course-v1:bvt+pp_01+2022' 'course-v1:deeptechforbusiness+FR+2021'
course_ids = [
  'course-v1:deeptechforbusiness+FR+2021', 'course-v1:deeptechforbusiness+EN+2021'
]

emails = sys.argv[1].split(";")

# One report every day + one report each month + a report after a year. 
# argv[2] should look like 'timePeriodToCheck;31'
daysLimit = int(sys.argv[2].split(";")[1])
all_users_data = {}
log.info('------------> Begin fetching user data and answers')

no_student = True

for course_id in course_ids:
  course_key = CourseLocator.from_string(course_id)
  course = get_course_by_id(course_key)
  course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)

  course_data = {}

  session = course_id.split('+')[1]
  
  for i in range(len(course_enrollments)):
    user = course_enrollments[i].user
    user_data = {}

    if str(user.email).find('@yopmail') != -1 or str(user.email).find('@weuplearning') != -1 or str(user.email).find(
            '@themoocagency') != -1:
      continue

    # FILTRER LES UTILISATEUR DU JOUR POUR RENDRE UN RAPPORT SANS ANCIENS UTILISATEURS :
    now = timezone.now()

    try:
      user_last_login = user.last_login
    except:
      continue

    if(user_last_login is None):
      continue

    if (now - timedelta(days=daysLimit) >= user_last_login):
      continue

    log.info('Treating --------> ' + str(user.email))
    no_student = False
    user_data["session"] = session

    from_zone = tz.gettz('UTC')
    to_zone = tz.gettz('Europe/Paris')
    utc = datetime.strptime(user.last_login.strftime('%d/%m/%Y %H:%M:%S'), '%d/%m/%Y %H:%M:%S')

    # Tell the datetime object that it's in UTC time zone since
    # datetime objects are 'naive' by default
    utc = utc.replace(tzinfo=from_zone)

    # Convert time zone
    '''local_time_actuastlitéle_login = utc.astimezone(to_zone)
    local_time_last_login = str(local_time_last_login).split('+')[0]

    user_data["last_login"] = local_time_last_login'''

    try:
      user_data["email"] = user.email
    except:
      try:
        user_data["email"] = json.loads(user.profile.custom_field)['email']
      except:
        user_data["email"] = 'n.a.'

    try:
      user_data["firstname"] = user.first_name.capitalize()
    except:
      try:
        user_data["firstname"] = json.loads(user.profile.custom_field)['first_name'].capitalize()
      except:
        user_data["firstname"] = 'n.a.'

    try:
      user_data["lastname"] = user.last_name.capitalize()
    except:
      try:
        user_data["lastname"] = json.loads(user.profile.custom_field)['last_name'].capitalize()
      except:
        user_data["lastname"] = 'n.a.'

    scorable_block_titles = []
    grading_context = grading_context_for_course(course)
    user_state_client = DjangoXBlockUserStateClient()
    list_question = []

    try:
      log.info(list(grading_context['all_graded_subsections_by_type'].keys()))
    
      exam_type = list(grading_context['all_graded_subsections_by_type'].keys())[0]
    except:
      log.info("user did not completed exams")
      continue

    for exam in list(grading_context['all_graded_subsections_by_type'].keys()):

      for section in grading_context['all_graded_subsections_by_type'][exam]: # QCM Intermédiaire
        scorable_block_titles = []
        for unit in section['scored_descendants']:
          scorable_block_titles.append((unit.location))
      k = 1
      list_question = []
      for block_location in scorable_block_titles:

        question = {}
        try:
          history_entries = list(user_state_client.get_history(user.username, block_location))
        except:
          question['choice'] = 'n.a.'
          continue

        problemNum = str(block_location)[-2:]
        question['problem'] = [problemNum, k]
        choices = []
        multiple_answer = ""

        if len(history_entries[0].state) == 3:
          question['choice'] = 'n.a.'
        else:
          try:
            for key, value in sorted(history_entries[0].state['student_answers'].items()):
              if len(history_entries[0].state['student_answers'].items()) > 1:
                multiple_answer += value + "\r\n"
                question['choice'] = multiple_answer
                log.info(multiple_answer)
              else:
                question['choice'] = value
          except:
            log.info("hello")
            continue

          correct = history_entries[0].state['correct_map'].items()
          correctness = ""
          for key, value in correct:
            if len(correct) > 1:
              correctness += value.get('correctness') + "\r\n"
              question['correctness'] = [correctness]
            else:
              correctness = value.get('correctness')
              question['correctness'] = [correctness]
        k += 1  

        list_question.append(question)

      data = {"general": user_data, 'list_question': list_question}
      course_data[str(user.id)+':'+exam] = data
      log.info("***************************************************")
      log.info(course_data)
      log.info("***************************************************")
    all_users_data[course_id] = course_data

log.info('------------> Finish fetching user data and answers')
log.info('------------> Begin write xlsx report')

# WRITE XLS
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= 'Rapport_deeptech'
filename = '/edx/app/edxapp/edx-themes/deeptechforbusiness/lms/static/utils/{}_deeptechforbusiness_grade_report.xlsx'.format(timestr)


headers = ['Adresse e-mail', 'Prénom', 'Nom', 'Session', 'Exam']

for i, header in enumerate(headers):
  sheet.cell(1, i+1, header)
j=2
for k, course_id in all_users_data.items():
  for key, user in course_id.items():
    sheet.cell(j, 1, user['general']['email'])
    sheet.cell(j, 2, user['general']['firstname'])
    sheet.cell(j, 3, user['general']['lastname'])
    sheet.cell(j, 4, user['general']['session'])
    sheet.cell(j, 5, key.split(':')[1])

    correctedExamGrade = 0
    i = 6
    for question in user['list_question']:
        try:
          sheet.cell(1, i, "question " + str(question['problem'][1]))
          correct = str(question['correctness'])
          sheet.cell(j, i, correct)
          sheet.cell(1, i+1, "choix " +  str(question['problem'][1]))
          choices = str(question['choice'])
          sheet.cell(j, i+1, choices)
        except:
          sheet.cell(1, i, "question " + str(question['problem'][1]))
          sheet.cell(j, i, "n.a.")
          sheet.cell(1, i+1, "choix " +  str(question['problem'][1]))
          sheet.cell(j, i+1, "n.a.")
        i +=2
    j += 1

#wb.save(filename)

# SEND MAILS
course_names = []
course_names_html = []
for course_id in course_ids: 
  course = get_course_by_id(CourseLocator.from_string(course_id)) 
  course_names.append(course.display_name_with_default)
  course_names_html.append("<li>"+ str(course.display_name_with_default)+"</li>")

output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ course_names_html +"<br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

for email in emails:
  if no_student :
    log.info('no_student')
    log.info(no_student)
    break

  part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
  fromaddr = "ne-pas-repondre@themoocagency.com"
  msg = MIMEMultipart()
  msg['From'] = fromaddr
  msg['To'] = email
  msg['Subject'] = "BVT_grade_report"
  attachment = _files_values
  part = MIMEBase('application', 'octet-stream')
  part.set_payload(attachment)
  encoders.encode_base64(part)
  part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
  msg.attach(part)
  server = smtplib.SMTP('mail3.themoocagency.com', 25)
  server.starttls()
  server.login('contact', 'waSwv6Eqer89')
  msg.attach(part2)
  text = msg.as_string()
  server.sendmail(fromaddr, email, text)
  server.quit()
  log.info('Email sent to '+str(email))

log.info('------------> Finish write xlsx report')
