# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()

#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################



from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from courseware.courses import get_course_by_id
from openedx.core.djangoapps.content.block_structure.api import get_course_in_cache
from completion.models import BlockCompletion

import json
import time
from openpyxl import Workbook

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


import logging
log = logging.getLogger()


emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")

all_users_data = {}
log.info('------------> Begin fetching user data and answers')

at_least_one_student = False
course_names = []
course_names_html = []


for course_id in course_ids:

    course_key = CourseLocator.from_string(course_id)
    course = get_course_by_id(course_key)
    course_names.append(course.display_name_with_default)
    course_names_html.append("<li>"+ str(course.display_name_with_default)+"</li>")

    course_structure = get_course_in_cache(course.id)

    # Create block_id_dict
    block_id_dict = {}
    id_in_chapter = []
    # initialize chapter var
    first = True
    unnecessary_chapter = False
    for section in course_structure:
        log.info(section)
        if str(section).find('chapter') != -1 :
            # remove last two sections :
            if str(section).find('259c99b5c0ba46318ca4a22f1d276380') != -1 or str(section).find('457a89c72983492cb08fc3beb1cc232f') != -1 :
                unnecessary_chapter = True
                continue
            # update object and then create the next chapter
            if first :
                chapter = str(section)
                first = False
            else:
                unnecessary_chapter = False
                block_id_dict[chapter] = id_in_chapter
                chapter = str(section)
                id_in_chapter = []
        elif str(section).find('html') != -1 or str(section).find('problem') != -1 :
            if not unnecessary_chapter :
                id_in_chapter.append(str(section).split("@")[2])
        else:
            continue

    course_data = {}

    course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)

    # for user in (list_of_student_modules):
    for i in range(len(course_enrollments)):
        user = course_enrollments[i].user

        # session
        session = course_id.split('+')[1]

        user_data = {}

        completed_block = []
        course_block_completions = BlockCompletion.get_learning_context_completions(user, course_key)

        for e, value in course_block_completions.items():
            completed_block.append(e.block_id)

        user_data["completion"] = completed_block
        
        at_least_one_student = True
        user_data["session"] = session

        try:
            user_data["email"] = user.email
        except:
            try:
                user_data["email"] = json.loads(user.profile.custom_field)['email']
            except:
                user_data["email"] = 'n.a.'

        if user_data["email"].find('@yopmail') != -1 or user_data["email"].find('@weuplearning') != -1 or user_data["email"].find('@themoocagency') != -1 or user_data["email"].find('@netexplo') != -1 :
            continue

        try:
            user_data["firstname"] = user.first_name.capitalize()
        except:
            try:
                user_data["firstname"] = json.loads(user.profile.custom_field)['first_name'].capitalize()
            except:
                user_data["firstname"] = 'n.a.'

        try:
            user_data["lastname"] = user.last_name.capitalize()
        except:
            try:
                user_data["lastname"] = json.loads(user.profile.custom_field)['last_name'].capitalize()
            except:
                user_data["lastname"] = 'n.a.'

        data = { "general": user_data }
        # data = { "general": user_data, 'list_question' : list_question }
        course_data[str(user.id)]= data
        all_users_data[course_id]= course_data

log.info('------------> Finish fetching user data and answers')

log.info('------------> Begin Calculate grades and write xlsx report')

# WRITE XLS
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= 'Rapport'
filename = '/home/edxtma/csv/{}_BMD_grade_report.xlsx'.format(timestr)

headers = ['Adresse e-mail', 'Prénom', 'Nom', 'Session', 'Chapitre 1', 'Chapitre 2', 'Chapitre 3', 'Chapitre 4', 'Chapitre 5', 'Chapitre 6', 'Chapitre 7', 'Chapitre 8', 'Chapitre 9', 'Chapitre 10', 'Chapitre 11', 'Chapitre 12', 'Chapitre 13', 'Chapitre 14', 'Chapitre 15' ]

for i, header in enumerate(headers):
    sheet.cell(1, i+1, header)
j=2
for k, course_id in all_users_data.items():

    for key, user in course_id.items():
        log.info(user['general']['email'])
        sheet.cell(j, 1, user['general']['email'])
        sheet.cell(j, 2, user['general']['firstname'])
        sheet.cell(j, 3, user['general']['lastname'])
        sheet.cell(j, 4, user['general']['session'])

        correctedExamGrade = 0
        i = 4
        for section, units in block_id_dict.items(): 
            log.info("---------------------------------")
            log.info(section)
            log.info(units)
            completion = 0
            block_in_section = 0

            for unit in units :
                block_in_section += 1 

                if unit in user['general']['completion'] :
                    completion += 1

            if block_in_section != 0:
                ratio = round(completion/block_in_section, 2)
                sheet.cell(j, i+1, ratio)
            i += 1
            
        j += 1


# SEND MAILS
output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ course_names_html +"<br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

for email in emails:
    if not at_least_one_student :
        log.info('at_least_one_student :')
        log.info(at_least_one_student)
        break

    part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "ne-pas-repondre@themoocagency.com"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = email
    msg['Subject'] = "BMD_grade_report"
    attachment = _files_values
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
    msg.attach(part)
    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, email, text)
    server.quit()
    log.info('Email sent to '+str(email))


log.info('------------> Finish calculate grades and write xlsx report')



# exemple Koa-prod
# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/bmd/lms/static/utils/completion_report.py 'eruch-ext@netexplo.org;lnyadanu@netexplo.org;melanie.zunino@weuplearning.com;cyril.adolf@weuplearning.com' 'course-v1:bmd+FR+2022_02_9-10'

