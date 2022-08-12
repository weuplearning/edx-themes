# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################


import json
import time
from openpyxl import Workbook
from datetime import datetime, date, timedelta
from django.utils import timezone
from dateutil import tz


from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from lms.djangoapps.courseware.courses import get_course_by_id
from lms.djangoapps.grades.context import grading_context_for_course
from lms.djangoapps.courseware.user_state_client import DjangoXBlockUserStateClient


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


import logging
log = logging.getLogger()


# course_ids is too long , do not use sys.argv[]
course_ids = [
  'course-v1:bvt+base_01+2022', 'course-v1:bvt+base_02+2022', 'course-v1:bvt+base_03+2022', 'course-v1:bvt+base_04+2022', 'course-v1:bvt+base_05+2022', 'course-v1:bvt+base_06+2022', 'course-v1:bvt+base_07+2022', 'course-v1:bvt+base_08+2022', 'course-v1:bvt+base_09+2022', 'course-v1:bvt+base_10+2022', 'course-v1:bvt+base_11+2022', 'course-v1:bvt+base_12+2022', 'course-v1:bvt+base_13+2022', 'course-v1:bvt+base_14+2022', 'course-v1:bvt+base_15+2022', 'course-v1:bvt+base_16+2022', 'course-v1:bvt+base_17+2022', 'course-v1:bvt+base_18+2022', 'course-v1:bvt+base_19+2022', 'course-v1:bvt+base_20+2022', 
  'course-v1:bvt+citernes_01+2022', 'course-v1:bvt+citernes_02+2022', 'course-v1:bvt+citernes_03+2022', 'course-v1:bvt+citernes_04+2022', 'course-v1:bvt+citernes_05+2022', 'course-v1:bvt+citernes_06+2022', 'course-v1:bvt+citernes_07+2022', 'course-v1:bvt+citernes_08+2022', 'course-v1:bvt+citernes_09+2022', 'course-v1:bvt+citernes_10+2022', 'course-v1:bvt+citernes_11+2022', 'course-v1:bvt+citernes_12+2022', 'course-v1:bvt+citernes_13+2022', 'course-v1:bvt+citernes_14+2022', 'course-v1:bvt+citernes_15+2022', 'course-v1:bvt+citernes_16+2022', 'course-v1:bvt+citernes_17+2022', 'course-v1:bvt+citernes_18+2022', 'course-v1:bvt+citernes_19+2022', 'course-v1:bvt+citernes_20+2022', 
  'course-v1:bvt+gpl_01+2022', 'course-v1:bvt+gpl_02+2022', 'course-v1:bvt+gpl_03+2022', 'course-v1:bvt+gpl_04+2022', 'course-v1:bvt+gpl_05+2022', 'course-v1:bvt+gpl_06+2022', 'course-v1:bvt+gpl_07+2022', 'course-v1:bvt+gpl_08+2022', 'course-v1:bvt+gpl_09+2022', 'course-v1:bvt+gpl_10+2022', 'course-v1:bvt+gpl_11+2022', 'course-v1:bvt+gpl_12+2022', 'course-v1:bvt+gpl_13+2022', 'course-v1:bvt+gpl_14+2022', 'course-v1:bvt+gpl_15+2022', 'course-v1:bvt+gpl_16+2022', 'course-v1:bvt+gpl_17+2022', 'course-v1:bvt+gpl_18+2022', 'course-v1:bvt+gpl_19+2022', 'course-v1:bvt+gpl_20+2022', 
  'course-v1:bvt+pp_01+2022', 'course-v1:bvt+pp_02+2022', 'course-v1:bvt+pp_03+2022', 'course-v1:bvt+pp_04+2022', 'course-v1:bvt+pp_05+2022', 'course-v1:bvt+pp_06+2022', 'course-v1:bvt+pp_07+2022', 'course-v1:bvt+pp_08+2022', 'course-v1:bvt+pp_09+2022', 'course-v1:bvt+pp_10+2022', 'course-v1:bvt+pp_11+2022', 'course-v1:bvt+pp_12+2022', 'course-v1:bvt+pp_13+2022', 'course-v1:bvt+pp_14+2022', 'course-v1:bvt+pp_15+2022', 'course-v1:bvt+pp_16+2022', 'course-v1:bvt+pp_17+2022', 'course-v1:bvt+pp_18+2022', 'course-v1:bvt+pp_19+2022', 'course-v1:bvt+pp_20+2022' 
]


emails = sys.argv[1].split(";")


# One report every day + one report each month + a report after a year. 
# argv[2] should look like 'timePeriodToCheck;31'
daysLimit = int(sys.argv[2].split(";")[1])


def updateGrade(problemNum, choices, answer_list):
  answered_true = 0
  grade = 0

  if choices == 'n.a.' or choices == []:
    return grade

  translated_list = []
  for choice in choices:
    index = choice.split('_')[1]
    translated_list.append(index)
  for choice in translated_list :
    problemRef = 'problem'+ problemNum 

    if choice in answer_list[problemRef]:
      answered_true += 1
    else:
      return grade

  if answered_true == len(answer_list[problemRef]):
    grade = 1
  else:
    grade = 0.5

  return grade



all_users_data = {}
log.info('------------> Begin fetching user data and answers')

no_student = True

for course_id in course_ids:
  course_key = CourseLocator.from_string(course_id)
  course = get_course_by_id(course_key)
  course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)
  course_name = course.display_name_with_default


  # DO NOT RENAME THE COURSE, IF NECESSARY, USE THE CONVERTER TO DO SO 
  # DO NOT RENAME THE COURSE, IF NECESSARY, USE THE CONVERTER TO DO SO 
  json_file_name = 'list_corrected_answer_' + str(course_name).replace(' ', '_') +'.json'
  with open('/edx/var/edxapp/media/microsites/bvt/answers_lists_files/'+json_file_name) as json_file:
    answer_list = json.load(json_file)
  # DO NOT RENAME THE COURSE, IF NECESSARY, USE THE CONVERTER TO DO SO 
  # DO NOT RENAME THE COURSE, IF NECESSARY, USE THE CONVERTER TO DO SO 


  course_data = {}

  # session
  session = course_id.split('+')[1]

  for i in range(len(course_enrollments)):
    user = course_enrollments[i].user
    user_data = {}

    if str(user.email).find('@yopmail') != -1 or str(user.email).find('@weuplearning') != -1 or str(user.email).find('@themoocagency') != -1 :
      continue

    # FILTRER LES UTILISATEUR DU JOUR POUR RENDRE UN RAPPORT SANS ANCIENS UTILISATEURS : 
    now = timezone.now()

    try:
      user_last_login = user.last_login
    except:
      continue

    if (now - timedelta(days=daysLimit) >= user_last_login ):
      continue

    log.info('Treating --------> ' + str(user.email))
    no_student = False
    user_data["session"] = session


    from_zone = tz.gettz('UTC')
    to_zone = tz.gettz('Europe/Paris')
    utc = datetime.strptime(user.last_login.strftime('%d/%m/%Y %H:%M:%S'), '%d/%m/%Y %H:%M:%S')

    # Tell the datetime object that it's in UTC time zone since 
    # datetime objects are 'naive' by default
    utc = utc.replace(tzinfo=from_zone)

    # Convert time zone
    local_time_actuastlitéle_login = utc.astimezone(to_zone)
    local_time_last_login = str(local_time_last_login).split('+')[0]

    user_data["last_login"] = local_time_last_login

    try:
      user_data["email"] = user.email
    except:
      try:
        user_data["email"] = json.loads(user.profile.custom_field)['email']
      except:
        user_data["email"] = 'n.a.'

    try:
      user_data["firstname"] = user.first_name.capitalize()
    except:
      try:
        user_data["firstname"] = json.loads(user.profile.custom_field)['first_name'].capitalize()
      except:
        user_data["firstname"] = 'n.a.'

    try:
      user_data["lastname"] = user.last_name.capitalize()
    except:
      try:
        user_data["lastname"] = json.loads(user.profile.custom_field)['last_name'].capitalize()
      except:
        user_data["lastname"] = 'n.a.'

    # Access Section
    scorable_block_titles = []
    grading_context = grading_context_for_course(course)
    user_state_client = DjangoXBlockUserStateClient()
    list_question = []

    for section in grading_context['all_graded_subsections_by_type']['Exam']:
      for unit in section['scored_descendants']:
        scorable_block_titles.append((unit.location))

    for block_location in scorable_block_titles:

      question = {}
      try:
        history_entries = list(user_state_client.get_history(user.username, block_location))
        log.info("block_location " + block_location)
        log.info("history_entries" + history_entries)
      except: 
        question['choice'] = 'n.a.'
        question['correctedGrade'] = 0
        question['time'] = 'n.a.'
        question['score'] = 'n.a.'
        continue

      problemNum = str(block_location)[-2:]
      question['problem'] = problemNum
      choices = []

      if len(history_entries[0].state) == 3 :
        question['choice'] = 'n.a.'
      else:
        for key, value in history_entries[0].state['student_answers'].items():
          choices = value
          question['choice'] = value

      # GRADE NEED TO BE RECALCULATE DUE TO BVT SCALE
      try:
        question['correctedGrade'] = updateGrade( problemNum, choices, answer_list)
      except:
        question['correctedGrade'] = 0

      try:
        question['time'] = history_entries[0].state['last_submission_time']
      except:
        question['time'] = 'n.a.'

      try:
        question['score'] = history_entries[0].state['score']['raw_earned']
      except:
        question['score'] = 'n.a.'

      list_question.append(question)

    data = { "general": user_data, 'list_question': list_question }
    course_data[str(user.id)]= data

  all_users_data[course_id]= course_data

log.info('------------> Finish fetching user data and answers')
log.info('------------> Begin Calculate grades and write xlsx report')

# WRITE XLS
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= 'Rapport'
filename = '/home/edxtma/csv/{}_BVT_grade_report.xlsx'.format(timestr)

headers = ['Adresse e-mail', 'Prénom', 'Nom', 'Session', 'Dernière connexion', 'Score' ,'Validation']

for i, header in enumerate(headers):
  sheet.cell(1, i+1, header)
j=2
for k, course_id in all_users_data.items():
  for key, user in course_id.items():
    sheet.cell(j, 1, user['general']['email'])
    sheet.cell(j, 2, user['general']['firstname'])
    sheet.cell(j, 3, user['general']['lastname'])
    sheet.cell(j, 4, user['general']['session'])
    sheet.cell(j, 5, user['general']['last_login'])

    correctedExamGrade = 0
    i = 6
    for question in user['list_question']:

      correctedGrade = question['correctedGrade']
      choices = ''
      for choice in question['choice'] :
        choices += str(choice) + ' '
      correctedExamGrade += int(correctedGrade)

    sheet.cell(j, i, correctedExamGrade)
    if correctedExamGrade >= 21: 
      sheet.cell(j, i+1, 'oui')
    else :
      sheet.cell(j, i+1, 'non')

    j += 1
sheet.cell(1, i+1, 'Note finale')


# SEND MAILS
course_names = []
course_names_html = []
for course_id in course_ids: 
  course = get_course_by_id(CourseLocator.from_string(course_id)) 
  course_names.append(course.display_name_with_default)
  course_names_html.append("<li>"+ str(course.display_name_with_default)+"</li>")

output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ course_names_html +"<br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

for email in emails:
  if no_student :
    log.info('no_student')
    log.info(no_student)
    break

  part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
  fromaddr = "ne-pas-repondre@themoocagency.com"
  msg = MIMEMultipart()
  msg['From'] = fromaddr
  msg['To'] = email
  msg['Subject'] = "BVT_grade_report"
  attachment = _files_values
  part = MIMEBase('application', 'octet-stream')
  part.set_payload(attachment)
  encoders.encode_base64(part)
  part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
  msg.attach(part)
  server = smtplib.SMTP('mail3.themoocagency.com', 25)
  server.starttls()
  server.login('contact', 'waSwv6Eqer89')
  msg.attach(part2)
  text = msg.as_string()
  server.sendmail(fromaddr, email, text)
  server.quit()
  log.info('Email sent to '+str(email))


log.info('------------> Finish calculate grades and write xlsx report')



# Exams occur everyday, send grade report after todays exam. Choose the right time to send a grade report in crontab. 

# 0 6 * * * /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/bvt/lms/static/utils/custom_grade_report_bvt.py 'cyril.adolf@weuplearning.com;alexandre.berteau@weuplearning.com' 'timePeriodToCheck;1'

# first day of every month at 6
# 0 6 1 * * /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/bvt/lms/static/utils/custom_grade_report_bvt.py 'cyril.adolf@weuplearning.com;alexandre.berteau@weuplearning.com' 'timePeriodToCheck;31'

# once a year the 1st of january at 6
# 0 6 1 1 * /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/bvt/lms/static/utils/custom_grade_report_bvt.py 'cyril.adolf@weuplearning.com;alexandre.berteau@weuplearning.com' 'timePeriodToCheck;365'
