# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import zipfile

import sys
importlib.reload(sys)

import os
from io import BytesIO

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################



import time
import json
import datetime

from opaque_keys.edx import locator

from opaque_keys.edx.locator import CourseLocator
from lms.djangoapps.courseware.courses import get_course_by_id
from student.models import CourseEnrollment

from lms.djangoapps.wul_apps.models import WulCourseEnrollment
from common.djangoapps.student.models import User, UserProfile
from lms.djangoapps.courseware.models import StudentModule
from opaque_keys.edx.keys import UsageKey

# entry point to the block_structure api.
from openedx.core.djangoapps.content.block_structure.api import get_course_in_cache

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


import logging
log = logging.getLogger()



try :
    emails = sys.argv[1].split(";")
except :
    emails = []


course_ids = [
    # "course-v1:hec-pole-emploi+WEB_1+2025",
    "course-v1:hec-pole-emploi+IP_1+2025",
    # "course-v1:hec-pole-emploi+IP_2+2025",
    # "course-v1:hec-pole-emploi+IP_3+2025",
    # "course-v1:hec-pole-emploi+IP_4+2025",
    "course-v1:hec-pole-emploi+NEG_1+2025",
    # "course-v1:hec-pole-emploi+NEG_2+2025",
    # "course-v1:hec-pole-emploi+NEG_3+2025",
    # "course-v1:hec-pole-emploi+NEG_4+2025",
]


courses_structure = {
    "course-v1:hec-pole-emploi+IP_1+2025" : [
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@28d184635ebe425cb5e026f2dfa0ec81",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@1a21a4b17e1b4b10be0ec6d0d671813c",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@9831a1a44c0f47d182ed63b8e5d63f07",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@43ac5cd8971e45628d1faedac3f13ac6",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@518c74129fdf4ab2ab0254b7b1c5b3ae",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@cd7b338899bc4d83b4b17ed78948bd96",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@7760f492dacf4954a230cb5db0a23fd1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@survey+block@e687558df0c0442c9bb8f3c04ac0791a",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@8fbdc506006b4308a2934e47d8824d6a",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@98378a6b5f3948c5aeb345b7fc50aafc",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@e970fa2deffc4fb29f2ea65599d44cae",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@e043ec09b1aa4c72a1cdd650532a34f6",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@d60be78923184a71b2833786e74b24f6",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@0aae72951c274598ab186b8f20b00a0f",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3dfb95aa6d164e74bbfbfc517db396da",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@5e7cc8bd6aff4a8cbbbaabf84c0be443",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@82720de74cb744a488b2e964094adcee",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@2edcdb853d6f43dca4036ed4726c18f0",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@277382d1c21a42ffa40920c5cebff66d",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@ac048762e124451597d2f877d3b36e6e",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@2f3e5b5a56094111acdaeb02b1c06692",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@306934e373284c3cad5d86763c608501",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@42896a1c9b204675a1d071540895292a",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@79d20703fa23446fb7b10abe9ca16d3c",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@53196212fb22480491860f41949f039d",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@0a11aea0f3de47498c98bf5715f92dd3",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@7ff88e1e618d4c8b8d90cfed4fdd1bfe",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3e2f4045813d4e848fe3c0e1f8d448b2",
        # module 2
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@c73d270134414e688bf87c3656962344",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3068a3072e0f452590557d78be443d2d",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@70e4109fbb114a129915c6e1b8251bae",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@163df922e8e84853891052c07e31204d",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@49498e6d3f5044bf9d2acafb5fe0e9c0",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@219ff84b743140e6a3ce77163d1931ea",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@7f87301e6313401d923e8e52e615ab23",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@c4da4aedec994c43ae0a7fe977b73703",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@2689deb47f3a4e83a1418126762e0753",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@f4b58e4417a34fd787bfec17c1b879a9",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@97e488b32edd401aa942e9b7a0f8fe17",
        # module2 session 2 
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@1117c041ada84d9886b84d5ff340ad04",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@4ae4e86c20e7455592003c89547a6530",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@7ea33b55afe1430eaf849849bb6fee76",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@87a97261cbf14a01bf201a091e5ad20f",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@b257a9215a664cf28ad27213ed22aa0c",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@49358d21bc6243a3aef484a4eadb22ff",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@13b711e81c804643be3022ca33104c46",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@e18195746b78407fb90c6cd0efc9af7d",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@e5cf9bf05bbc41e080066906237c7ce4",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@abdcbe5eb98244bab36fc38745e33b6e",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@19445afb73974819a47197dbece02063",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@fda8d18a99a04a11be30c6664abae4de",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3987744b9bc6479ca16c59c6bab5e46a",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@aca7f3e1f8ae43d9a6f4abf310ce4e5d",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@aeca70fb5c1a4963ad60a420e22be9bb",
        # MODULE 3 S1
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@9f82416fa7ad4a52b667f074c12a5e1b",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@e601cefcfdbf4f0aae31644b961d3917",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@cbd333c8abb744f3a21fcb2ee240251a",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@88c6adadfedb41b1b1ef9bbc71f91cb9",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@73e0d82563564eadb01458f87948bed9",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@054f39c3045c40dc92b67e6130a4d413",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@a36f1a2f07af46f7807bbae6015af051",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@d5eb105cd46c4737889dc2826955a86b",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@8c2de07783f243b5be81fae01347642c",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@1a5990751a034603bede59b9dee8de90",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@d39e9a989a6c43fb99b26400cfa2d735",

        # MODULE 3 S2
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@66c02803a94c49b097c07948a651efe4",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@70aece870caf4e6e865f4b6ea7b6fd6a",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3ef6a513dcef460bae19d5998388158c",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@4051e64c94f246ba9444235df9ab43a1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@6afe9c6819e0472da2e66a728304cdb3",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@50215a3202ea49b1b87ed497bed45ed6",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@ebf66cc8ba1f4bfaa05ac00d84305fb5",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@abfa11adc53e4effb41d02cb703c4c1d",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@82293e6eae1b43d3a68bb5f20fe11e10",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@7fd0e4f4eae44cdb9ef97a4eada9ed4c",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@c9f12b60a62540b9867807b399933b22",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@a627883894b2477aa5a74c0c54feee6f",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@7671ab77639c48d1bd90b6a5ac3bc60f",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@7c45fc1a25a5449f815930a303ed7fff",
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@932fe8cb8a42408aa1da00410537a021",
        # EVALUATION FINALE
        "block-v1:hec-pole-emploi+IP_1+2025+type@problem+block@80fda803382f4890a6cefe849ee354b2"        
    ],
    # "course-v1:hec-pole-emploi+IP_2+2025" : {
    #     "block-v1:hec-pole-emploi+IP_1+2025+type@scorm+block@76f94648789249ae8be25ee7b6b5e61d": "Scorm module"
    # },
    # "course-v1:hec-pole-emploi+IP_3+2025" : {
    #     "block-v1:hec-pole-emploi+IP_1+2025+type@scorm+block@76f94648789249ae8be25ee7b6b5e61d": "Scorm module"
    # },
    # "course-v1:hec-pole-emploi+IP_4+2025" : {
    #     "block-v1:hec-pole-emploi+IP_1+2025+type@scorm+block@76f94648789249ae8be25ee7b6b5e61d": "Scorm module"
    # },
    "course-v1:hec-pole-emploi+NEG_1+2025" : [
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@7e4621359ba24e3d83a4fe54c71c47dd",
        # Session 1
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@685b7601be0d491ebead11f173957850",
        "espaceur car le bloc précédent contient 2 réponses",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@06b3b8fa36954ae29ae550895486fa73",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@scorm+block@825663373a73460bb0c7bc9f9feead19",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@4c27cb93579e411a9e62d4a8eebe9d08",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@6ce2c8d97115492eb18b2020a61afc85",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@478f768fde1c446d90620417879dca66",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@732b176b20a64dfb92bd61e4b203585a",
        # "block-v1:hec-pole-emploi+NEG_1+2025+type@invideoquiz+block@ed7d523977b8495aa79cdf085c2c0bbc",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@0860e52e39c5425cb3a64c04edd38c57",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@c6a73372690a4ef3a4b05ac1e8d88b67",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@900d75ce7588400e98beb7025b12bae5",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@152c31e325b949f4b35c2dc4bd4c7407",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@3b696510868443fe824a38268c0894c5",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@9de12455cd4f43f2b126d98685191f46",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@e6cbb477f25249eb8b92aba139a8745a",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@survey+block@cfca8207bff94103adfc2f7bb6beaa40",

        # Session 2
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@5a0e70d1eec94f4197f4c5b4459a4559",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@scorm+block@bc27da548b35437cb6988e887f797a6e",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@05e6336409e0486a8f736532e3ec41a4",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@7dcefa2bef8547d7a5b95aceb575a0a6",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@15b3601c830c453f94c150f63e37e5ad",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@79a8b86323584384be0802f7a39123ea",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@problem+block@c3b71a3c543645968ffe7cf2c047ff5d",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@f4fb777804544404ae6023fbef4c747b",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@9e7bf750d57242faa271a75c41a7efc5",
    ],
    # "course-v1:hec-pole-emploi+NEG_2+2025" : {
    #     "block-v1:hec-pole-emploi+NEG_2+2025+type@scorm+block@9f693653691f4b5daa0de77dd18d3e12": "Scorm module"
    # },
    # "course-v1:hec-pole-emploi+NEG_3+2025" : {
    #     "block-v1:hec-pole-emploi+NEG+2023+type@scorm+block@75d793663c1d4256aedd2ae063d3c978": "Scorm module"
    # },
    # "course-v1:hec-pole-emploi+NEG_4+2025" : {
    #     "block-v1:hec-pole-emploi+NEG+2023+type@scorm+block@75d793663c1d4256aedd2ae063d3c978": "Scorm module"
    # },
    # "course-v1:hec-pole-emploi+WEB_1+2025" : {},
    # "course-v1:hec-pole-emploi+WEB_2+2025" : {},
    # "course-v1:hec-pole-emploi+WEB_3+2025" : {},
    # "course-v1:hec-pole-emploi+WEB_4+2025": {}
}
listed_answer_list ={
    "course-v1:hec-pole-emploi+NEG_1+2025" : [
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@7e4621359ba24e3d83a4fe54c71c47dd",
        "685b7601be0d491ebead11f173957850_3_1",
        "685b7601be0d491ebead11f173957850_2_1",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@06b3b8fa36954ae29ae550895486fa73",
        # "01bc47db4ca041938955462ec6d490c5_2_1",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@4c27cb93579e411a9e62d4a8eebe9d08",
        "6ce2c8d97115492eb18b2020a61afc85_2_1",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@478f768fde1c446d90620417879dca66",
        "732b176b20a64dfb92bd61e4b203585a_2_1",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@0860e52e39c5425cb3a64c04edd38c57",
        "c6a73372690a4ef3a4b05ac1e8d88b67_2_1",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@900d75ce7588400e98beb7025b12bae5",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@152c31e325b949f4b35c2dc4bd4c7407",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@3b696510868443fe824a38268c0894c5",
        "9de12455cd4f43f2b126d98685191f46_2_1",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@e6cbb477f25249eb8b92aba139a8745a",
        "enjoy",
        "recommend",
        "learn",
        "1758925220212",
        "1758925284278",
        "1758925285193",
        "1758925286127",
        "1758925286960",
        "1758925288444",
        "1758925357444",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@5a0e70d1eec94f4197f4c5b4459a4559",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@05e6336409e0486a8f736532e3ec41a4",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@f4fb777804544404ae6023fbef4c747b",
        "block-v1:hec-pole-emploi+NEG_1+2025+type@video+block@9e7bf750d57242faa271a75c41a7efc5"
    ],
    "course-v1:hec-pole-emploi+IP_1+2025" : [
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@28d184635ebe425cb5e026f2dfa0ec81",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@1a21a4b17e1b4b10be0ec6d0d671813c",
        "9831a1a44c0f47d182ed63b8e5d63f07_2_1",
        "43ac5cd8971e45628d1faedac3f13ac6_2_1",
        "518c74129fdf4ab2ab0254b7b1c5b3ae_2_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@cd7b338899bc4d83b4b17ed78948bd96",
        "7760f492dacf4954a230cb5db0a23fd1_2_1",
        # "block-v1:hec-pole-emploi+IP_1+2025+type@survey+block@e687558df0c0442c9bb8f3c04ac0791a",
        # 24 questions
        "1761642289653",
        "1761642296044",
        "1761642387848",
        "1761642401688",
        "1761642424920",
        "1761642439935",
        "1761642836487",
        "1761642850831",
        "1761642863591",
        "1761642875838",
        "1761642890510",
        "1761642905677",
        "1761642918581",
        "1761642933972",
        "1761642944611",
        "1761642955051",
        "1761642966323",
        "1761642979570",
        "1761642990394",
        "1761643004345",
        "1761643013977",
        "1761643022616",
        "1761643030688",
        "1761643039536",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@8fbdc506006b4308a2934e47d8824d6a",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@98378a6b5f3948c5aeb345b7fc50aafc",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@e970fa2deffc4fb29f2ea65599d44cae",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@e043ec09b1aa4c72a1cdd650532a34f6",
        "d60be78923184a71b2833786e74b24f6_2_1",
        "0aae72951c274598ab186b8f20b00a0f_2_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3dfb95aa6d164e74bbfbfc517db396da",
        "5e7cc8bd6aff4a8cbbbaabf84c0be443_2_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@82720de74cb744a488b2e964094adcee",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@2edcdb853d6f43dca4036ed4726c18f0",
        "277382d1c21a42ffa40920c5cebff66d_2_1",
        "277382d1c21a42ffa40920c5cebff66d_2_2",
        "277382d1c21a42ffa40920c5cebff66d_2_3",
        "277382d1c21a42ffa40920c5cebff66d_2_4",
        "277382d1c21a42ffa40920c5cebff66d_2_5",
        "277382d1c21a42ffa40920c5cebff66d_2_6",
        "277382d1c21a42ffa40920c5cebff66d_2_7",
        "277382d1c21a42ffa40920c5cebff66d_2_8",
        "ac048762e124451597d2f877d3b36e6e_2_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@2f3e5b5a56094111acdaeb02b1c06692",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@306934e373284c3cad5d86763c608501",
        "42896a1c9b204675a1d071540895292a_2_1",
        "79d20703fa23446fb7b10abe9ca16d3c_2_1",
        "53196212fb22480491860f41949f039d_2_1",
        "0a11aea0f3de47498c98bf5715f92dd3_2_1",
        "0a11aea0f3de47498c98bf5715f92dd3_3_1",
        "0a11aea0f3de47498c98bf5715f92dd3_4_1",
        "0a11aea0f3de47498c98bf5715f92dd3_5_1",
        "7ff88e1e618d4c8b8d90cfed4fdd1bfe_2_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3e2f4045813d4e848fe3c0e1f8d448b2",
        # module 2
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@c73d270134414e688bf87c3656962344",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3068a3072e0f452590557d78be443d2d",
        "70e4109fbb114a129915c6e1b8251bae_2_1",
        "70e4109fbb114a129915c6e1b8251bae_2_2",
        "70e4109fbb114a129915c6e1b8251bae_2_3",
        "70e4109fbb114a129915c6e1b8251bae_2_4",
        "70e4109fbb114a129915c6e1b8251bae_2_5",
        "70e4109fbb114a129915c6e1b8251bae_2_6",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@163df922e8e84853891052c07e31204d",
        "49498e6d3f5044bf9d2acafb5fe0e9c0_2_1",
        "49498e6d3f5044bf9d2acafb5fe0e9c0_2_2",
        "49498e6d3f5044bf9d2acafb5fe0e9c0_2_3",
        "49498e6d3f5044bf9d2acafb5fe0e9c0_2_4",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@219ff84b743140e6a3ce77163d1931ea",
        "7f87301e6313401d923e8e52e615ab23_2_1",
        "7f87301e6313401d923e8e52e615ab23_2_2",
        "7f87301e6313401d923e8e52e615ab23_2_3",
        "7f87301e6313401d923e8e52e615ab23_2_4",
        "7f87301e6313401d923e8e52e615ab23_2_5",
        "7f87301e6313401d923e8e52e615ab23_2_6",
        "7f87301e6313401d923e8e52e615ab23_2_7",
        "7f87301e6313401d923e8e52e615ab23_2_8",
        "7f87301e6313401d923e8e52e615ab23_2_9",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@c4da4aedec994c43ae0a7fe977b73703",
        "2689deb47f3a4e83a1418126762e0753_2_1",
        "2689deb47f3a4e83a1418126762e0753_2_2",
        "2689deb47f3a4e83a1418126762e0753_2_3",
        "2689deb47f3a4e83a1418126762e0753_2_4",
        "2689deb47f3a4e83a1418126762e0753_2_5",
        "f4b58e4417a34fd787bfec17c1b879a9_2_1",
        "f4b58e4417a34fd787bfec17c1b879a9_3_1",
        "f4b58e4417a34fd787bfec17c1b879a9_4_1",
        "f4b58e4417a34fd787bfec17c1b879a9_5_1",
        "f4b58e4417a34fd787bfec17c1b879a9_6_1",
        "f4b58e4417a34fd787bfec17c1b879a9_7_1",
        "f4b58e4417a34fd787bfec17c1b879a9_8_1",
        "f4b58e4417a34fd787bfec17c1b879a9_9_1",
        "f4b58e4417a34fd787bfec17c1b879a9_10_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@97e488b32edd401aa942e9b7a0f8fe17",
        # module2 session 2 
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@1117c041ada84d9886b84d5ff340ad04",
        "4ae4e86c20e7455592003c89547a6530_2_1",
        "4ae4e86c20e7455592003c89547a6530_3_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@7ea33b55afe1430eaf849849bb6fee76",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@87a97261cbf14a01bf201a091e5ad20f",
        "b257a9215a664cf28ad27213ed22aa0c_2_1",
        "b257a9215a664cf28ad27213ed22aa0c_3_1",
        "b257a9215a664cf28ad27213ed22aa0c_4_1",
        "b257a9215a664cf28ad27213ed22aa0c_5_1",
        "b257a9215a664cf28ad27213ed22aa0c_6_1",
        "b257a9215a664cf28ad27213ed22aa0c_7_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@49358d21bc6243a3aef484a4eadb22ff",
        "13b711e81c804643be3022ca33104c46_2_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@e18195746b78407fb90c6cd0efc9af7d",
        "e5cf9bf05bbc41e080066906237c7ce4_2_1",
        "abdcbe5eb98244bab36fc38745e33b6e_2_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@19445afb73974819a47197dbece02063",
        "fda8d18a99a04a11be30c6664abae4de_2_1",
        "fda8d18a99a04a11be30c6664abae4de_3_1",
        "fda8d18a99a04a11be30c6664abae4de_4_1",
        "fda8d18a99a04a11be30c6664abae4de_5_1",
        "fda8d18a99a04a11be30c6664abae4de_6_1",
        "fda8d18a99a04a11be30c6664abae4de_7_1",
        "fda8d18a99a04a11be30c6664abae4de_8_1",
        "fda8d18a99a04a11be30c6664abae4de_9_1",
        "fda8d18a99a04a11be30c6664abae4de_10_1",
        "fda8d18a99a04a11be30c6664abae4de_11_1",
        "fda8d18a99a04a11be30c6664abae4de_12_1",
        "fda8d18a99a04a11be30c6664abae4de_13_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3987744b9bc6479ca16c59c6bab5e46a",
        "aca7f3e1f8ae43d9a6f4abf310ce4e5d_2_1",
        "aca7f3e1f8ae43d9a6f4abf310ce4e5d_3_1",
        "aca7f3e1f8ae43d9a6f4abf310ce4e5d_4_1",
        "aca7f3e1f8ae43d9a6f4abf310ce4e5d_5_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@aeca70fb5c1a4963ad60a420e22be9bb",
        # MODULE 3 S1
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@9f82416fa7ad4a52b667f074c12a5e1b",
        "e601cefcfdbf4f0aae31644b961d3917_2_1",
        "e601cefcfdbf4f0aae31644b961d3917_3_1",
        "e601cefcfdbf4f0aae31644b961d3917_4_1",
        "e601cefcfdbf4f0aae31644b961d3917_5_1",
        "e601cefcfdbf4f0aae31644b961d3917_6_1",
        "e601cefcfdbf4f0aae31644b961d3917_7_1",
        "e601cefcfdbf4f0aae31644b961d3917_8_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@cbd333c8abb744f3a21fcb2ee240251a",
        "88c6adadfedb41b1b1ef9bbc71f91cb9_2_1",
        "88c6adadfedb41b1b1ef9bbc71f91cb9_3_1",
        "88c6adadfedb41b1b1ef9bbc71f91cb9_4_1",
        "73e0d82563564eadb01458f87948bed9_2_1",
        "73e0d82563564eadb01458f87948bed9_3_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@054f39c3045c40dc92b67e6130a4d413",
        "a36f1a2f07af46f7807bbae6015af051_2_1",
        "a36f1a2f07af46f7807bbae6015af051_3_1",
        "a36f1a2f07af46f7807bbae6015af051_4_1",
        "a36f1a2f07af46f7807bbae6015af051_5_1",
        "d5eb105cd46c4737889dc2826955a86b_2_1",
        "d5eb105cd46c4737889dc2826955a86b_3_1",
        "d5eb105cd46c4737889dc2826955a86b_4_1",
        "d5eb105cd46c4737889dc2826955a86b_5_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@8c2de07783f243b5be81fae01347642c",
        "1a5990751a034603bede59b9dee8de90_2_1",
        "1a5990751a034603bede59b9dee8de90_3_1",
        "1a5990751a034603bede59b9dee8de90_4_1",
        "1a5990751a034603bede59b9dee8de90_5_1",
        "1a5990751a034603bede59b9dee8de90_6_1",
        "1a5990751a034603bede59b9dee8de90_7_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@d39e9a989a6c43fb99b26400cfa2d735",
        # MODULE 3 S2
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@66c02803a94c49b097c07948a651efe4",
        "70aece870caf4e6e865f4b6ea7b6fd6a_2_1",
        "70aece870caf4e6e865f4b6ea7b6fd6a_3_1",
        "70aece870caf4e6e865f4b6ea7b6fd6a_4_1",
        "70aece870caf4e6e865f4b6ea7b6fd6a_5_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@3ef6a513dcef460bae19d5998388158c",
        "4051e64c94f246ba9444235df9ab43a1_2_1",
        "4051e64c94f246ba9444235df9ab43a1_3_1",
        "4051e64c94f246ba9444235df9ab43a1_4_1",
        "6afe9c6819e0472da2e66a728304cdb3_2_1",
        "6afe9c6819e0472da2e66a728304cdb3_3_1",
        "6afe9c6819e0472da2e66a728304cdb3_4_1",
        "6afe9c6819e0472da2e66a728304cdb3_5_1",
        "6afe9c6819e0472da2e66a728304cdb3_6_1",
        "6afe9c6819e0472da2e66a728304cdb3_7_1",
        "6afe9c6819e0472da2e66a728304cdb3_8_1",
        "6afe9c6819e0472da2e66a728304cdb3_9_1",
        "6afe9c6819e0472da2e66a728304cdb3_10_1",
        "6afe9c6819e0472da2e66a728304cdb3_11_1",
        "6afe9c6819e0472da2e66a728304cdb3_12_1",
        "6afe9c6819e0472da2e66a728304cdb3_13_1",
        "6afe9c6819e0472da2e66a728304cdb3_14_1",
        "6afe9c6819e0472da2e66a728304cdb3_15_1",
        "6afe9c6819e0472da2e66a728304cdb3_16_1",
        "6afe9c6819e0472da2e66a728304cdb3_17_1",
        "6afe9c6819e0472da2e66a728304cdb3_18_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@50215a3202ea49b1b87ed497bed45ed6",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@ebf66cc8ba1f4bfaa05ac00d84305fb5",
        "abfa11adc53e4effb41d02cb703c4c1d_2_1",
        "abfa11adc53e4effb41d02cb703c4c1d_3_1",
        "abfa11adc53e4effb41d02cb703c4c1d_4_1",
        "abfa11adc53e4effb41d02cb703c4c1d_5_1",
        "abfa11adc53e4effb41d02cb703c4c1d_6_1",
        "abfa11adc53e4effb41d02cb703c4c1d_7_1",
        "abfa11adc53e4effb41d02cb703c4c1d_8_1",
        "abfa11adc53e4effb41d02cb703c4c1d_9_1",
        "abfa11adc53e4effb41d02cb703c4c1d_10_1",
        "abfa11adc53e4effb41d02cb703c4c1d_11_1",
        "abfa11adc53e4effb41d02cb703c4c1d_12_1",
        "abfa11adc53e4effb41d02cb703c4c1d_13_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@82293e6eae1b43d3a68bb5f20fe11e10",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_2_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_3_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_4_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_5_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_6_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_7_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_8_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_9_1",
        "7fd0e4f4eae44cdb9ef97a4eada9ed4c_10_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@c9f12b60a62540b9867807b399933b22",
        "a627883894b2477aa5a74c0c54feee6f_2_1",
        "a627883894b2477aa5a74c0c54feee6f_3_1",
        "a627883894b2477aa5a74c0c54feee6f_4_1",
        "a627883894b2477aa5a74c0c54feee6f_5_1",
        "a627883894b2477aa5a74c0c54feee6f_6_1",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@7671ab77639c48d1bd90b6a5ac3bc60f",
        "block-v1:hec-pole-emploi+IP_1+2025+type@video+block@7c45fc1a25a5449f815930a303ed7fff",
        "932fe8cb8a42408aa1da00410537a021_2_1",
        "932fe8cb8a42408aa1da00410537a021_3_1",
        "932fe8cb8a42408aa1da00410537a021_4_1",
        # EVALUATION FINALE
        "80fda803382f4890a6cefe849ee354b2_2_1",
        "80fda803382f4890a6cefe849ee354b2_3_1",
        "80fda803382f4890a6cefe849ee354b2_4_1",
        "80fda803382f4890a6cefe849ee354b2_5_1",
        "80fda803382f4890a6cefe849ee354b2_6_1",
        "80fda803382f4890a6cefe849ee354b2_7_1",
        "80fda803382f4890a6cefe849ee354b2_8_1",
        "80fda803382f4890a6cefe849ee354b2_9_1",
        "80fda803382f4890a6cefe849ee354b2_10_1",
        "80fda803382f4890a6cefe849ee354b2_11_1"
    ]
}

translate_answer_dict = {
    "course-v1:hec-pole-emploi+NEG_1+2025" : {
        # Premiere offre - quiz
        "6ce2c8d97115492eb18b2020a61afc85" : {
            "choice_0" : "Quand vous êtes l'acheteur",
            "choice_1" : "Quand vous êtes le vendeur",
            "choice_2" : "Lorsque vous avez peu d'informations sur la valeur de ce que vous négociez",
            "choice_3" : "Lorsque vous avez collecté beaucoup d'informations sur la valeur de ce que vous négociez",
            "choice_4" : "Lorsque qu'il y a d'autres offres concurrentes"
        },
        # Prix de réserve - quiz
        "732b176b20a64dfb92bd61e4b203585a" : {
            "choice_0" : "Je divulgue mon prix de réserve au début de la négociation pour instaurer un climat de confiance",
            "choice_1" : "Je ne divulgue pas mon prix de réserve",
            "choice_2" : "Je divulgue mon prix de réserve à la fin de la négociation pour pousser l'autre partie à conclure un accord"
        },
        # Zone de négociation - quiz
        "c6a73372690a4ef3a4b05ac1e8d88b67" : {
            "choice_0" : "Abandonner la négociation sans conclure d'accord",
            "choice_1" : "En tant que vendeur, descendre en dessous de votre prix de réserve pour ouvrir la zone de négociation",
            "choice_2" : "Rendre l'autre partie consciente de la zone de négociation négative et la contourner."
        },
        # Le BATNA - quiz
        "9de12455cd4f43f2b126d98685191f46" : {
            "choice_0" : "Votre BATNA peut être légèrement différente de votre prix de réserve",
            "choice_1" : "Votre BATNA fait référence au nombre d’offres alternatives que vous avez.",
            "choice_2" : "Plus votre BATNA est intéressante, plus vous aurez de pouvoir dans la négociation",
            "choice_3" : "Il est essential d’essayer d’améliorer vos options alternatives en amont d’une négociation (afin d’augmenter votre BATNA)",
            "choice_4" : "En révélant votre BATNA à votre interlocuteur, vous augmentez votre levier dans la négociation"
        }
    },
    "course-v1:hec-pole-emploi+IP_1+2025" : {
        # Quiz 1 - module 1
        "9831a1a44c0f47d182ed63b8e5d63f07" : {
            "choice_0" : "A : J'aurais tenté l’aventure entrepreneuriale cette idée d'entreprise. Pas de gloire sans risques !",
            "choice_1" : "B : Je me serais contenté de mon travail. Un tiens vaut mieux que deux tu l’auras…, on ne sait jamais ce qui pourrait arriver en repartant de zéro !"
        },
        "43ac5cd8971e45628d1faedac3f13ac6" : {
            "choice_0" : "A : Il semble y avoir quelque chose qui cloche avec cette idée d'entreprise. Si possible, j'aurais repris mon ancien emploi, et j'aurais patienté en attendant de trouver une meilleure idée.",
            "choice_1" : "B : Il semble y avoir quelques problèmes avec l'idée. J'aurais essayé de découvrir l'origine de ces problèmes et j'aurais cherché à développer des plans B ainsi qu’à adapter et affiner la stratégie commerciale. J'aurais persévéré pour atteindre mon objectif !"
        },
        "518c74129fdf4ab2ab0254b7b1c5b3ae" : {
            "choice_0" : "A : Je ne pense pas que cela soit utile. Il est difficile de prédire à quoi ressemblera l'avenir, je préfère donc me concentrer sur l'ici et maintenant de l'entreprise.",
            "choice_1" : "B : Je pense que c'est une bonne idée. Le fait de réfléchir au futur lorsqu'on entreprend un projet commercial se traduira par un avantage concurrentiel à long terme."
        },
        "0aae72951c274598ab186b8f20b00a0f" : {
            "choice_0" : "Ne gaspillez pas votre énergie, attendez de devoir réagir à un problème !",
            "choice_1" : "Agissez d'abord au lieu de suivre les autres !",
            "choice_2" : "Ne vous contentez pas de réagir aux changements de l'environnement, changez activement votre environnement !",
            "choice_3" : "Acceptez que parfois vous ne pouvez pas changer des circonstances défavorables !",
            "choice_4" : "Commencez une action vous-même, n'attendez pas de devoir réagir !",
            "choice_5" : "Changez les circonstances défavorables au lieu d'attendre que quelqu'un d'autre le fasse !",
            "choice_6" : "Soyez différent des autres, ne vous contentez pas de les copier mais réfléchissez à la façon dont vous pouvez vous différencier et mettre en valeur votre entreprise !",
            "choice_7" : "Concentrez-vous sur ce que les autres font bien, et essayez de les imiter !",
            "choice_8" : "Ne laissez pas vos idées rester des pensées et des rêves, réalisez-les !",
            "choice_9" : "Occupez-vous de ce qui est le plus important maintenant, les rêves et les objectifs peuvent attendre !"
        },
        "13b711e81c804643be3022ca33104c46" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "f4b58e4417a34fd787bfec17c1b879a9_3_1" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "f4b58e4417a34fd787bfec17c1b879a9_4_1" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "f4b58e4417a34fd787bfec17c1b879a9_5_1" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "f4b58e4417a34fd787bfec17c1b879a9_6_1" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "f4b58e4417a34fd787bfec17c1b879a9_7_1" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "f4b58e4417a34fd787bfec17c1b879a9_8_1" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "f4b58e4417a34fd787bfec17c1b879a9_9_1" : {
            "choice_0" : "Oui",
            "choice_1" : "Non"
        },
        "88c6adadfedb41b1b1ef9bbc71f91cb9_3_1" : {
            "choice_0" : "1",
            "choice_1" : "2",
            "choice_2" : "3",
            "choice_3" : "4",
            "choice_4" : "5",
            "choice_5" : "6",
            "choice_6" : "7",
            "choice_7" : "8",
            "choice_8" : "9",
            "choice_9" : "10"
        },
        "80fda803382f4890a6cefe849ee354b2_2_1" : {
            "choice_0" : "Être persévérant",
            "choice_1" : "Être proactif",
            "choice_2" : "Être compétent",
            "choice_3" : "Être orienté vers le futur"
        },
        "80fda803382f4890a6cefe849ee354b2_3_1" : {
            "choice_0" : "Prendre des initiatives et agir sans attendre des instructions",
            "choice_1" : "Attendre qu'on vous dise quoi faire",
            "choice_2" : "Éviter les responsabilités",
            "choice_3" : "Ne réagir que lorsque les problèmes apparaissent"
        },
        "80fda803382f4890a6cefe849ee354b2_4_1" : {
            "choice_0" : "Voir les problèmes comme une occasion d'apprendre et de vous améliorer",
            "choice_1" : "Essayer de vous concentrer sur des tâches faciles",
            "choice_2" : "Ralentir ou reporter votre objectif",
            "choice_3" : "Compter principalement sur les autres"
        },
        "80fda803382f4890a6cefe849ee354b2_5_1" : {
            "choice_0" : "L'expérimentation n'est pas nécessaire si je crois déjà que mon idée est bonne",
            "choice_1" : "Les expériences peuvent aider à tester l'idée commerciale à petite échelle",
            "choice_2" : "Les expériences sont utiles car les erreurs m'aident à apprendre",
            "choice_3" : "Les expériences peuvent m'aider à tester une idée rapidement"
        },
        "80fda803382f4890a6cefe849ee354b2_6_1" : {
            "choice_0" : "Technique Pourquoi-Alors",
            "choice_1" : "Technique Si-Comment",
            "choice_2" : "Technique Si-Alors",
            "choice_3" : "Technique Quand-Comment"
        },
        "80fda803382f4890a6cefe849ee354b2_7_1" : {
            "choice_0" : "Demander aux gens quels sont leurs besoins ou problèmes, et réfléchir à des solutions",
            "choice_1" : "Copier des idées d'entreprises prospères à l'étranger",
            "choice_2" : "Vous fier uniquement à vos compétences actuelles",
            "choice_3" : "Attendre que les clients vous disent ce qu'ils veulent"
        },
        "80fda803382f4890a6cefe849ee354b2_8_1" : {
            "choice_0" : "L'objectif doit être décrit clairement",
            "choice_1" : "L'objectif doit être mon objectif personnel",
            "choice_2" : "L'objectif doit correspondre exactement à mon type d'entreprise",
            "choice_3" : "Je dois déjà avoir un plan détaillé"
        },
        "80fda803382f4890a6cefe849ee354b2_9_1" : {
            "choice_0" : "Vous devriez toujours avoir des plans B",
            "choice_1" : "Vous devriez partir d'un objectif SMART PI",
            "choice_2" : "Vous devriez d'abord planifier vos actions et seulement ensuite penser à vos ressources",
            "choice_3" : "Après la planification, vous devriez vérifier régulièrement vos progrès"
        },
        "80fda803382f4890a6cefe849ee354b2_10_1" : {
            "choice_0" : "Les considérer comme une occasion d'apprendre et de vous améliorer",
            "choice_1" : "Essayer de les ignorer",
            "choice_2" : "Vous sentir découragé et arrêter d'essayer",
            "choice_3" : "Les cacher aux autres"
        },
        "80fda803382f4890a6cefe849ee354b2_11_1" : {
            "choice_0" : "Utiliser différentes sources de retours inhabituelles",
            "choice_1" : "Ne demander qu'à une seule personne",
            "choice_2" : "Éviter les retours négatifs",
            "choice_3" : "Attendre que les autres vous donnent des retours"
        }
    }
}




users_per_course = dict()
# list_chapters_name = dict()
all_user_set = set()
# list_of_student_scorms = list()
videos_list = list()


# Modifier si plusieurs version pour chaque theme
def course_name(course_id):
    if course_id.find('IP_') != -1 : 
        return "Initiative Personnelle (IP)"
    elif course_id.find('WEB_') != -1 :
        return "Webinaire (WEB)"
    else : 
        return "Négociation (NEG)"



for course_id in course_ids:

    course_key = CourseLocator.from_string(course_id)
    course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)
    users_data = dict()

    for i in range(len(course_enrollments)):
        user = course_enrollments[i].user


        # Escape fake email address
        # if str(user.email).find('cyril.adolf@weuplearning.com')  == -1 :
        if user.email.find("@example")!= -1 or user.email.find("@themoocagency") != -1 or user.email.find("@fake")!= -1:
        # if user.email.find("@example")!= -1 or user.email.find("@themoocagency") != -1 or user.email.find("@weuplearning")!= -1 or user.email.find("@fake")!= -1:
            continue


        user_data = dict()

        user_data.update({
            "id": getattr(user, "id", ""),
            "username": getattr(user, "username", ""),
            "email": getattr(user, "email", ""),
            "date_joined": getattr(user, "date_joined", "").strftime('%Y-%m-%d %H:%M:%S') if getattr(user, "date_joined", None) else "",
            "last_login": getattr(user, "last_login", "").strftime('%Y-%m-%d %H:%M:%S') if getattr(user, "last_login", None) else "",
            "name": user.profile.name
        })


        user_row = []
        video_dict = dict()


        log.info('treating user :')
        log.info(user.email)

        course_key = locator.CourseLocator.from_string(str(course_id))
        collected_block_structure = get_course_in_cache(course_key)


        try : 
            user_problems = StudentModule.objects.filter(student=user, course_id__exact=course_id, module_type="problem").values('state','module_state_key')
            user_videos = StudentModule.objects.filter(student=user, course_id__exact=course_id, module_type="video").values('state','module_state_key')
            user_surveys = StudentModule.objects.filter(student=user, course_id__exact=course_id, module_type="survey").values('state','module_state_key')


            answer_dict = {}

            for unit in courses_structure[course_id] :

                log.info('unit')
                log.info(unit)

                if unit.find('problem') != -1 :
                    log.info('"unit problem"')
                    log.info(unit)

                    for user_problem in user_problems:

                        if str(user_problem['module_state_key']) != str(unit) :
                            continue

                        log.info('user_problem')
                        log.info(user_problem['state'])
                        log.info(user_problem['module_state_key'])

                        if 'student_answers' not in user_problem['state'] :
                            continue

                        json_state = json.loads(user_problem['state'])


                        block_id = str(user_problem['module_state_key']).split('+block@')[1]
                        log.info('block_id !!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                        log.info(block_id)

                        if block_id in translate_answer_dict[course_id] :
                            for key, value in json_state['student_answers'].items() :


                                log.info(value)
                                if isinstance(value , list) :

                                    allChoices = ''
                                    for choice in value :
                                        log.info('choice before translation')
                                        log.info(choice)


                                        if choice in translate_answer_dict[course_id][block_id] :
                                            translated_value = translate_answer_dict[course_id][block_id][choice]
                                            log.info('translated_value')
                                            log.info(translated_value)
                                            allChoices += translated_value + ' ; '

                                    answer_dict[key] = allChoices
                                    break
                                else :
                                    log.info(key)
                                    log.info(value)
                                    log.info(translate_answer_dict[course_id][block_id][value])

                                    answer_dict[key] = translate_answer_dict[course_id][block_id][value]
                                    break

                            #     if key in translate_answer_dict[course_id][block_id] :
                            #         translated_value = translate_answer_dict[course_id][block_id][key]
                            #         log.info('translated_value')
                            #         log.info(translated_value)
                            #         answer_dict[key] = str(translated_value)
                            # break

                            log.info('before if isinstance ')
                        elif isinstance(json_state['student_answers'] , dict) :
                            for key, value in json_state['student_answers'].items() :
                                log.info(key)
                                # vérifier si key dans translate_answer_dict
                                if key in translate_answer_dict[course_id] :
                                    if value in translate_answer_dict[course_id][key] :
                                        translated_value = translate_answer_dict[course_id][key][value]
                                        log.info('translated_value')
                                        log.info(translated_value)
                                        answer_dict[key] = str(translated_value)
                                    else :
                                        answer_dict[key] = str(value)
                                else :
                                    answer_dict[key] = str(value)
                            break

                        else :
                            answer_dict[str(user_problem['module_state_key'])] = json_state['student_answers']
                            break



                if unit.find('video') != -1 :
                    log.info('"unit video"')
                    log.info(unit)
                    for user_video in user_videos:
                        if str(user_video['module_state_key']) != str(unit) :
                            continue
                        json_state = json.loads(user_video['state'])
                        log.info("json_state")
                        log.info(json_state)
                        answer_dict[str(user_video['module_state_key'])] = json_state['saved_video_position']
                        break


                if unit.find('survey') != -1 :
                    log.info('"unit survey"')
                    log.info(unit)

                    for user_survey in user_surveys:
                        if str(user_survey['module_state_key']) != str(unit) :
                            continue
                        json_state = json.loads(user_survey['state'])
                        log.info("json_state survey")
                        log.info(json_state)

                        if isinstance(json_state['choices'] , dict) :
                            for key, value in json_state['choices'].items() :

                                if course_id == "course-v1:hec-pole-emploi+NEG_1+2025":
                                    if value == 'Y' :
                                        answer_dict[key] = str("Vrai")
                                    else :
                                        answer_dict[key] = str("Faux")
                                elif course_id == "course-v1:hec-pole-emploi+IP_1+2025":
                                    
                                    log.info(key)
                                    log.info(value)

                                    if value == 'Y' :
                                        answer_dict[key] = str("Pas du tout d'accord")
                                    elif value == 'N' :
                                        answer_dict[key] = str("Pas d'accord")
                                    elif value == 'M' :
                                        answer_dict[key] = str("Plutôt pas d'accord")
                                    elif value == '1761642164633' :
                                        answer_dict[key] = str("Neutre")
                                    elif value == '1761642216223' :
                                        answer_dict[key] = str("Plutôt d'accord")
                                    elif value == '1761642217999' :
                                        answer_dict[key] = str("D'accord")
                                    else :
                                        answer_dict[key] = str("Tout à fait d'accord")
                            break

                        answer_dict[str(user_survey['module_state_key'])] = json_state['completed']
                        break


        except : 
            log.info('no student module found')
            log.info('no student module found')
            log.info('no student module found')


        log.info('@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ answer_dict')
        log.info(answer_dict)


        try:
            wul_course_enrollment = WulCourseEnrollment.objects.get(course_enrollment_edx__user=user, course_enrollment_edx__course_id=course_key)
            global_time_tracking = wul_course_enrollment.global_time_tracking
        except:
            global_time_tracking = 0

        user_data["global_time_tracking"] = datetime.timedelta(seconds=global_time_tracking)



        user_row = [user_data["username"],user_data["email"],user_data["name"],user_data["date_joined"],user_data["last_login"], user_data["global_time_tracking"]]

        log.info('keyDict utility')
        for unit in listed_answer_list[course_id] :
            log.info('unit keyDict')
            log.info(unit)
            if unit in answer_dict :

                log.info(answer_dict[unit])
                user_row.append(answer_dict[unit])
            else :
                user_row.append('')


        log.info('user_row')
        log.info(user_row)

        users_data[user.username.capitalize()] = user_row
    users_per_course[course_id] = users_data


## Workbook
wb = Workbook()
wb.remove(wb.active)


def create_sheet_function(course_id, users, workbook):

    sheet_name = course_name(course_id)
    common_header = ["Username","Email","Nom complet","Date de création de compte","Date de dernière connexion","Temps passé total"] 

    if course_id == "course-v1:hec-pole-emploi+NEG_1+2025" :
        questions_header = [
            "introduction - video", 

            # session 1
            "A votre avis 1/2", 
            "A votre avis 2/2", 
            "Le cas: Ancienne papeterie - video",
            # "Le cas: Ancienne papeterie - checkbox",
            # "Simulation: Ancienne papeterie - scorm", 
            "La première offre - video", 
            "Quiz: La première offre?", 
            "Prix de réserve - video",
            "Prix de réserve - quiz",
            "Zone de négociation - video",
            "Zone de négociation - quiz",
            "Le biais du coût irrécupérable - video",
            "Q & R Le prix de réserve - video",
            "Le BATNA - video",
            "Le BATNA - quiz",
            "La conclusion - video",
            "La conclusion - survey 1/10",
            "La conclusion - survey 2/10",
            "La conclusion - survey 3/10",
            "La conclusion - survey 4/10",
            "La conclusion - survey 5/10",
            "La conclusion - survey 6/10",
            "La conclusion - survey 7/10",
            "La conclusion - survey 8/10",
            "La conclusion - survey 9/10",
            "La conclusion - survey 10/10",

            # session 2
            "Le cas : l'atelier - video",
            # scorm
            "L'ancrage - video",
            "Point d'aspiration - video",
            "Établir des relations - video"
        ]
    elif course_id == "course-v1:hec-pole-emploi+IP_1+2025" :
        questions_header = [
            # Module 1 s1
            "Introduction - video",
            "Études de cas : Élodie - video",
            "Q 1 - quiz",
            "Q 2 - quiz",
            "Q 3 - quiz",
            "Persévérance entrepreneuriale - video",
            "Q 1 - quiz",
            "Auto-évaluation - survey 1/24",
            "Auto-évaluation - survey 2/24",
            "Auto-évaluation - survey 3/24",
            "Auto-évaluation - survey 4/24",
            "Auto-évaluation - survey 5/24",
            "Auto-évaluation - survey 6/24",
            "Auto-évaluation - survey 7/24",
            "Auto-évaluation - survey 8/24",
            "Auto-évaluation - survey 9/24",
            "Auto-évaluation - survey 10/24",
            "Auto-évaluation - survey 11/24",
            "Auto-évaluation - survey 12/24",
            "Auto-évaluation - survey 13/24",
            "Auto-évaluation - survey 14/24",
            "Auto-évaluation - survey 15/24",
            "Auto-évaluation - survey 16/24",
            "Auto-évaluation - survey 17/24",
            "Auto-évaluation - survey 18/24",
            "Auto-évaluation - survey 19/24",
            "Auto-évaluation - survey 20/24",
            "Auto-évaluation - survey 21/24",
            "Auto-évaluation - survey 22/24",
            "Auto-évaluation - survey 23/24",
            "Auto-évaluation - survey 24/24",
            "L'autonomie - video",
            "Etude de cas d'autonomie - video",
            "Conclusions et Home Challenge - video",
            # Module 1 s2
            "L'introduction session 2 - video",
            "Q 1 - quiz",
            "Révision du home challenge - quiz", 
            "Déclenchement Individuel : Passage à l'Action - video",
            "Q 1 - quiz",
            "Introduction à l'orientation vers le futur - video",
            "Étude de cas orientation future - video",
            "Exercice: Le futur soi - Opportunité 1 - quiz",
            "Exercice: Le futur soi - Action pour opportunité 1 - quiz", 
            "Exercice: Le futur soi - Opportunité 2 - quiz",
            "Exercice: Le futur soi - Action pour opportunité 2 - quiz", 
            "Exercice: Le futur soi - Problème 1 - quiz",
            "Exercice: Le futur soi - Action pour problème 1 - quiz", 
            "Exercice: Le futur soi - Problème 2 - quiz",
            "Exercice: Le futur soi - Action pour problème 2 - quiz", 
            "Étape 3 : Routine d'orientation future - quiz",
            "Introduction à la persévérance - video",
            "Etude de cas sur la persévérance - video",
            "Q 1 - quiz",
            "Q 2 - quiz",
            "Étape 1 : Identifier le scénario catastrophe - quiz",
            "Étape 2 - 1 : Stratégies de gestion - quiz", 
            "Étape 2 - 2 : Stratégies de gestion - quiz", 
            "Étape 2 - 3 : Stratégies de gestion - quiz", 
            "Étape 2 - 4 : Stratégies de gestion - quiz", 
            "Étape 3 : Visualisation et action - quiz",
            "Récapitulation session 2 - video",
            # Module 2 s1
            "Introduction module 2 - video",
            "Principes d'action pour l'identification des opportunités - video",
            "Exercice : Changements technologiques - quiz", 
            "Exercice : Problèmes/opportunités découlant des changements technologiques - quiz", 
            "Exercice : Changements démographiques et sociaux - quiz", 
            "Exercice : Problèmes/opportunités découlant des changements démographiques et sociaux - quiz", 
            "Exercice : Autres changements (par exemple, changements réglementaires) - quiz", 
            "Exercice : Problèmes découlant d'autres changements - quiz", 
            "Débriefing et deuxième partie de « Changements dans votre environnement » - video",
            "Exercice : Changement 1 - quiz", 
            "Exercice : Concept 1 - quiz", 
            "Exercice : Changement 2 - quiz", 
            "Exercice : Concept 2 - quiz", 
            "Création d'opportunités - video",
            "Exercice: SCAMPER - produit service - quiz", 
            "Exercice: SCAMPER - Composantes - quiz", 
            "Exercice: SCAMPER - Substituer - quiz", 
            "Exercice: SCAMPER - Combiner - quiz", 
            "Exercice: SCAMPER - Adapter - quiz", 
            "Exercice: SCAMPER - Modifier - quiz", 
            "Exercice: SCAMPER - Produire - quiz", 
            "Exercice: SCAMPER - Eliminer - quiz", 
            "Exercice: SCAMPER - Renverser - quiz", 
            "Débriefing et définition des objectifs - video",
            "Q 1 - S - quiz",
            "Q 1 - M - quiz",
            "Q 1 - A - quiz",
            "Q 1 - R - quiz",
            "Q 1 - T - quiz",
            "Exercice: définition des objectifs - Objectif - quiz", 
            "Exercice: définition des objectifs - Spécifique - quiz", 
            "Exercice: définition des objectifs - Mesurable - quiz", 
            "Exercice: définition des objectifs - Ambitieux - quiz", 
            "Exercice: définition des objectifs - Réaliste - quiz", 
            "Exercice: définition des objectifs - Temporellement défini - quiz", 
            "Exercice: définition des objectifs - Introduit un élément nouveau - quiz", 
            "Exercice: définition des objectifs - Tient compte des tendances futures ? - quiz", 
            "Exercice: définition des objectifs - Reformulation de votre objectif - quiz", 
            "Récapitulatif et défi à domicile - video",
            # Module 2 s2
            "Introduction - video",
            "Fixation d'objectif SMART-IP - Objectif - quiz", 
            "Fixation d'objectif SMART-IP - Vérifier - quiz", 
            "Planification financière - video",
            "Sources conventionnelles de financement et auto-évaluation financière - video",
            "Exercise: votre projet personnel - Prêts bancaires - quiz", 
            "Exercise: votre projet personnel - Prêts gouvernementaux - quiz", 
            "Exercise: votre projet personnel - Capital risque - quiz", 
            "Exercise: votre projet personnel - Business Angels - quiz", 
            "Exercise: votre projet personnel - Subventions - quiz", 
            "Exercise: votre projet personnel - Crowdfunding - quiz", 
            "Amorçage financier - video",
            "Le concept d'amorçage financier - quiz", 
            "Conférence sur l'amorçage financier - video",
            "Q 1 - quiz",
            "Exercice : amorçage financier - quiz",
            "Plan d'action - video",
            "Exercice: plan d'action — partie 1 - Matériel/outils - quiz",
            "Exercice: plan d'action — partie 1 - Ou puis-je me le procurer - quiz",
            "Exercice: plan d'action — partie 1 - Main-d'oeuvre - quiz",
            "Exercice: plan d'action — partie 1 - Ou puis-je me le procurer - quiz",
            "Exercice: plan d'action — partie 1 - Temps personnel à consacrer - quiz",
            "Exercice: plan d'action — partie 1 - Ou puis-je me le procurer - quiz",
            "Exercice: plan d'action — partie 1 - Information - quiz",
            "Exercice: plan d'action — partie 1 - Ou puis-je me le procurer - quiz",
            "Exercice: plan d'action — partie 1 - Argent - quiz",
            "Exercice: plan d'action — partie 1 - Ou puis-je me le procurer - quiz",
            "Exercice: plan d'action — partie 1 - Autre - quiz",
            "Exercice: plan d'action — partie 1 - Ou puis-je me le procurer - quiz",
            "Plan d'action — partie 2 - video",
            "Exercice: plan d'action — partie 2 - Important, mais non urgent - quiz",
            "Exercice: plan d'action — partie 2 - Important et urgent - quiz",
            "Exercice: plan d'action — partie 2 - Non important et non urgent - quiz",
            "Exercice: plan d'action — partie 2 - Non important, mais urgent - quiz",
            "Récapitulatif et défi à domicile - video",
            # Module 3 S1
            "Introduction module 3 - video",
            "Le suivi des exercices à la maison - Si 1 - quiz",
            "Le suivi des exercices à la maison - Alors 1 - quiz",
            "Le suivi des exercices à la maison - Si 2 - quiz",
            "Le suivi des exercices à la maison - Alors 2 - quiz",
            "Le suivi des exercices à la maison - Si 3 - quiz",
            "Le suivi des exercices à la maison - Alors 3 - quiz",
            "Le suivi des exercices à la maison - Première action - quiz",
            "La mise en œuvre du plan d’action et l’importance des erreurs - video",
            "Q 1 - sentiment erreur - quiz",
            "Q 1 - note erreur - quiz",
            "Q 1 - considération erreur - quiz",
            "Exercice : Réflexion sur les erreurs et les leçons à tirer - Etape 1 - quiz", 
            "Exercice : Réflexion sur les erreurs et les leçons à tirer - Etape 2 - quiz", 
            "Expérience portant sur l’entreprise de Mathieu - video",
            "Quiz - Qui - quiz",
            "Quiz - Quoi - quiz",
            "Quiz - Comment - quiz",
            "Quiz - Lecon - quiz",
            "Exercice: Expérience individuelle - Idee - quiz", 
            "Exercice: Expérience individuelle - Hypothèse - quiz", 
            "Exercice: Expérience individuelle - Description - quiz", 
            "Exercice: Expérience individuelle - Seuil - quiz", 
            "Sources de feedback et feedback négatif - video",
            "Quiz - Source 1 - quiz",
            "Quiz - Idée 1 - quiz",
            "Quiz - Source 2 - quiz",
            "Quiz - Idée 2 - quiz",
            "Quiz - Source 3 - quiz",
            "Quiz - Idée 3 - quiz",
            "Récapitulatif et exercice à faire à la maison - video",
            # Module 3 S2
            "Introduction - video",
            "Exercice : Feedback et leçons tirées - Source inhabituelle - quiz", 
            "Exercice : Feedback et leçons tirées - Aide - quiz", 
            "Exercice : Feedback et leçons tirées - Feedback négatif - quiz", 
            "Exercice : Feedback et leçons tirées - Leçons - quiz", 
            "Rappel des principes d’action de l’IP - video",
            "Quiz - Le D... - quiz",
            "Quiz - Comportement o... - quiz",
            "Quiz - Comportement p... - quiz",
            "Classification des Principes d'Action Entrepreneuriale - 1 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 2 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 3 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 4 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 5 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 6 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 7 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 8 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 9 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 10 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 11 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 12 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 13 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 14 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 15 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 16 - quiz", 
            "Classification des Principes d'Action Entrepreneuriale - 17 - quiz", 
            "Introduction de l’étude de cas - video",
            "Travail d’étude de cas partie 1 - video",
            "Questions - Partie 1 - Outils - quiz",  
            "Questions - Partie 1 - Idée 1 - quiz",  
            "Questions - Partie 1 - Idée 2 - quiz",  
            "Questions - Partie 1 - Outil utilisé - quiz",  
            "Questions - Partie 2 - Objectif commercial - quiz",  
            "Questions - Partie 2 - Spécifique - quiz",  
            "Questions - Partie 2 - Mesurable - quiz",  
            "Questions - Partie 2 - Ambitieux - quiz",  
            "Questions - Partie 2 - Réaliste - quiz",  
            "Questions - Partie 2 - Terminé - quiz",  
            "Questions - Partie 2 - Novateur - quiz",  
            "Questions - Partie 2 - Tendances futures - quiz",  
            "Travail d’étude de cas partie 2 - video",
            "Exercice : Planification Commerciale pour Karine - Matériels/outils - quiz",  
            "Exercice : Planification Commerciale pour Karine - Main-d'oeuvre - quiz",  
            "Exercice : Planification Commerciale pour Karine - Temps à consacrer - quiz",  
            "Exercice : Planification Commerciale pour Karine - Information - quiz",  
            "Exercice : Planification Commerciale pour Karine - Autre - quiz",  
            "Exercice : Planification Commerciale pour Karine - Argent - quiz",  
            "Exercice : Planification Commerciale pour Karine - Stratégie de financement - quiz",  
            "Exercice : Planification Commerciale pour Karine - Plan d'action - quiz",  
            "Exercice : Planification Commerciale pour Karine - Techniques spécifiques - quiz",  
            "Travail d’étude de cas partie 3 - video",
            "Exercice : Feedback et Expérimentation pour Karine - Recherche et feedback - quiz",
            "Exercice : Feedback et Expérimentation pour Karine - Idée à tester - quiz",
            "Exercice : Feedback et Expérimentation pour Karine - Hypothèse - quiz",
            "Exercice : Feedback et Expérimentation pour Karine - Description - quiz",
            "Exercice : Feedback et Expérimentation pour Karine - Seuil de validation - quiz",
            "Projet personnel - video",
            "Contrat avec soi-même - video", 
            "Contrat d'Engagement - Date - quiz",
            "Contrat d'Engagement - Prenom - quiz",
            "Contrat d'Engagement - Signature - quiz",
            # Final evaluation
            "Évaluation finale - Lequel des énoncés suivants ne fait PAS partie de l'initiative personnelle ? - quiz",
            "Évaluation finale - Lequel des comportements suivants est un exemple de proactivité ? - quiz",
            "Évaluation finale - Quel principe d'action devriez-vous suivre si vous voulez devenir plus persévérant ? - quiz",
            "Évaluation finale - Quelle affirmation concernant le test d'une idée commerciale avec des expériences est incorrecte ? - quiz",
            "Évaluation finale - Quelle technique vous aide à préparer vos problèmes et actions futurs ? - quiz",
            "Évaluation finale - Lequel des énoncés suivants est une bonne façon de découvrir de nouvelles opportunités commerciales ? - quiz",
            "Évaluation finale - Que signifie le \"S\" dans l'objectif \"SMART-IP\" (\"Spécifique\") ? - quiz",
            "Évaluation finale - Quelle affirmation concernant la planification des actions est incorrecte ? - quiz",
            "Évaluation finale - Quelle est une bonne façon de gérer les erreurs dans votre entreprise ? - quiz",
            "Évaluation finale - Lequel des énoncés suivants est un principe d'action utile pour rechercher des retours ? - quiz"
        ]

    common_header = common_header + questions_header


    sheet = workbook.create_sheet(sheet_name)

    for i, header in enumerate(common_header):
        sheet.cell(row=1, column=(i+1)).value = header

    j=2
    for user in users:
        user_row = user[1]
        l=0
        for value in user_row :
            sheet.cell(row=j, column=(l+1)).value = value
            l=l+1
        j=j+1



for course_id in course_ids:
    users_data = users_per_course[course_id]
    if users_data == []:
        continue
    ordered_users = sorted(users_data.items(), key=lambda x: x[1])
    create_sheet_function(course_id, ordered_users, wb)



filename = "hec_grade_report.xlsx"
filepath = '/edx/var/edxapp/media/microsites/hec-pole-emploi/reports/{}'.format(filename)
wb.save(filepath)
output = BytesIO()
_files_values = output.getvalue()
html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de donn&eacute;es HEC France-Travail</p></body></html>"

## Send email
for email in emails:
    part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "HEC France-Travail <ne-pas-repondre@themoocagency.com>"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = email
    msg['Subject'] = "Data report HEC France-Travail"

    attachment = _files_values

    with open(filepath, 'rb') as f:
        attachment = f.read()

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= {}".format(filename))
    msg.attach(part)

    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, email, text)
    server.quit()

    log.info('Email sent to '+email)




# sudo /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/hec-pole-emploi/lms/utils/data_report_hec.py "cyril.adolf@weuplearning.com"

