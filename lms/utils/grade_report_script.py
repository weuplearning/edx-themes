# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################


import json
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from lms.djangoapps.courseware.courses import get_course_by_id


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


import logging
log = logging.getLogger()


emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")

all_users_data = {}
sections_id = {
  "course-v1:arif+lsfin+fr": [
    "e96db9cf67814538bc5daeb77b3b0942",
    "04c1034316e54202ba25e295fad1ed21",
    "a2f4f7aa8364457ca76766719d805c98",
    "0f9a55284f4d4ae385be87dd177b2bc9",
    "07e5c4cef83341cb852bee4005a97eb6",
    "376428f5beeb48a18814c7acf9189fc0",
    "4bc6ac548f3a43199e6605c2ca666cc9"
  ],
  "course-v1:arif+lsfin+en": [
    "e96db9cf67814538bc5daeb77b3b0942",
    "04c1034316e54202ba25e295fad1ed21",
    "a2f4f7aa8364457ca76766719d805c98",
    "0f9a55284f4d4ae385be87dd177b2bc9",
    "07e5c4cef83341cb852bee4005a97eb6",
    "376428f5beeb48a18814c7acf9189fc0",
    "4bc6ac548f3a43199e6605c2ca666cc9"
  ]
}




for course_id in course_ids:
  course_key = CourseLocator.from_string(course_id)
  course = get_course_by_id(course_key)
  course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)
  course_name = course.display_name_with_default

  course_data = {}

  for i in range(len(course_enrollments)):
    user = course_enrollments[i].user
    user_data = {}

    enrollment = course_enrollments[i]
    if str(user.email).find('@yopmail') != -1 or str(user.email).find('@weuplearning') != -1 or str(user.email).find('@themoocagency') != -1 :
      continue


    try:
      user_data["email"] = user.email
    except:
      try:
        user_data["email"] = json.loads(user.profile.custom_field)['email']
      except:
        user_data["email"] = 'n.a.'
    
    user_data["name"] = user.profile.name


    try:
      last_login = user.last_login.strftime("%d/%m/%Y")
    except:
      last_login = None

    try:
      date_joined = user.date_joined.strftime("%d/%m/%Y")
    except:
      date_joined = None


    user_data["register_date"] = date_joined
    user_data["last_login"] = last_login

    # Grade
    user_cf = json.loads(user.profile.custom_field) 
    user_grade = []
    for section_id in sections_id[course_id]:
      if section_id in user_cf:
        user_grade.append(user_cf[section_id])
      else: 
        user_grade.append('n.a.')

    user_data['grade'] = user_grade


    try :
        certificate = user_cf["success_date_"+ course_id]
    except :
        certificate = ''
    user_data['certif'] = certificate

    data = { "general": user_data }
    course_data[str(user.id)]= data

  all_users_data[course_id]= course_data

# log.info('------------> Finish fetching user data and answers')
# log.info('------------> Begin Calculate grades and write xlsx report')

# WRITE XLS
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= 'Rapport de notes'
filename = '/home/edxtma/csv/{}_arif_grade_report.xlsx'.format(timestr)

headers = ['Email', 'Nom d\'utilisateur' , 'Date d\'inscription','Date de dernière connexion', 'Note section 1', 'Note section 2', 'Note section 3', 'Note section 4', 'Note section 5', 'Note section 6', 'Note section 7', 'Date de validation']
for i, header in enumerate(headers):
  sheet.cell(1, i+1, header)
  sheet.cell(1, i+1).fill = PatternFill("solid", fgColor="1D2235")
  sheet.cell(1, i+1).font = Font(b=False, color="FFFFFF")

j=2
for k, course_id in all_users_data.items():

  j+=1
  sheet.cell(j, 1, k)
  sheet.cell(j, 1).fill = PatternFill("solid", fgColor="1D2235")
  sheet.cell(j, 1).font = Font(b=False, color="FFFFFF")
  j+=1
  for key, user in course_id.items():

    sheet.cell(j, 1, user['general']['email'])
    sheet.cell(j, 2, user['general']['name'] )
    sheet.cell(j, 3, user['general']['register_date'])
    sheet.cell(j, 4, user['general']['last_login'])

    sheet.cell(j, 5, user['general']['grade'][0])
    sheet.cell(j, 6, user['general']['grade'][1])
    sheet.cell(j, 7, user['general']['grade'][2])
    sheet.cell(j, 8, user['general']['grade'][3])
    sheet.cell(j, 9, user['general']['grade'][4])
    sheet.cell(j, 10, user['general']['grade'][5])
    sheet.cell(j, 11, user['general']['grade'][6])

    sheet.cell(j, 12, user['general']['certif'])

    j += 1


# SEND MAILS
# course_names = []
course_names_html = []
for course_id in course_ids: 
  course = get_course_by_id(CourseLocator.from_string(course_id)) 
  # course_names.append(course.display_name_with_default)
  course_names_html.append("<li>"+ str(course.display_name_with_default)+"</li>")

output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ course_names_html +"<br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

for email in emails:

  part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
  fromaddr = "ne-pas-repondre@themoocagency.com"
  msg = MIMEMultipart()
  msg['From'] = fromaddr
  msg['To'] = email
  msg['Subject'] = "arif_grade_report"
  attachment = _files_values
  part = MIMEBase('application', 'octet-stream')
  part.set_payload(attachment)
  encoders.encode_base64(part)
  part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
  msg.attach(part)
  server = smtplib.SMTP('mail3.themoocagency.com', 25)
  server.starttls()
  server.login('contact', 'waSwv6Eqer89')
  msg.attach(part2)
  text = msg.as_string()
  server.sendmail(fromaddr, email, text)
  server.quit()
  log.info('Email sent to '+str(email))


log.info('------------> Finish calculate grades and write xlsx report')


# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/arif/lms/utils/grade_report_script.py 'cyril.adolf@weuplearning.com' 'course-v1:arif+lsfin+en;course-v1:arif+lsfin+fr'


