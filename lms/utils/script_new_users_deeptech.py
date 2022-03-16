# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()

#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^
#                START BEYOND THIS LINE
#############################################################################################################################


from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from courseware.courses import get_course_by_id
from lms.djangoapps.grades.context import grading_context_for_course
from lms.djangoapps.courseware.user_state_client import DjangoXBlockUserStateClient
from lms.djangoapps.wul_apps.models import WulCourseEnrollment


import json
import time
from openpyxl import Workbook

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import logging
log = logging.getLogger()




emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")




all_users_data = {}
log.info('------------> Begin fetching user data and answers')
at_least_one_student = False


for course_id in course_ids:
  course_key = CourseLocator.from_string(course_id)
  course = get_course_by_id(course_key)
  course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)

  for i in range(len(course_enrollments)):
    user = course_enrollments[i].user
    log.info(user)
    user_data = {}

    tma_enrollment,is_exist=WulCourseEnrollment.objects.get_or_create(course_enrollment_edx=course_enrollments[i])

    if user.email.find('@weuplearning') != -1 or user.email.find('@yopmail') != -1 or user.email.find('@the-mooc-agency') != -1:
      log.info(user.email)
      log.info('test user')
      continue

    # Hide 
    # ONLY SAVE IF USER IS NEW (J-30) 
    string_data = str(course_enrollments[i])
    date_registration = datetime.strptime(string_data.split(' ')[3].replace('(',''), '%Y-%m-%d')

    today = datetime.now()
    test_substract = (today - date_registration).days


    if test_substract > 31 :
      log.info(user.email)
      log.info(test_substract)
      continue
    # Dev Cyril End


    bugged = ['bvt10112_encRg','bvt1011_t9M4C', 'alex_staff']
    if str(user) in bugged:
      continue
    
    at_least_one_student = True
    # Update object with user data without grades
    try:
      user_data["username"] = user.username
    except:
      try:
        user_data["username"] = json.loads(user.profile.custom_field)['username']
      except:
        user_data["username"] = 'n.a.'

    try:
      user_data["firstname"] = user.first_name.capitalize()
    except:
      try:
        user_data["firstname"] = json.loads(user.profile.custom_field)['firstname'].capitalize()
      except:
        user_data["firstname"] = 'n.a.'

    try:
      user_data["lastname"] = user.last_name.capitalize()
    except:
      try:
        user_data["lastname"] = json.loads(user.profile.custom_field)['lastname'].capitalize()
      except:
        user_data["lastname"] = 'n.a.'

    # Time tracking
    try:
      seconds = tma_enrollment.global_time_tracking
      minute = seconds // 60
      time_tracking = int(minute)
    except:
      time_tracking = int(0)

    user_data['time_tracking'] = time_tracking

    data = { "general": user_data }
    all_users_data[str(user.id)]= data

log.info('------------> Finish fetching user data and answers')

log.info('------------> Begin Calculate grades and write xlsx report')
# Grades need to be recalculate :
# We need to generate un json when the course is converted. 
# Then we can fetch it using the course name (value given when creating the tar.gz) 

course_names = []
course_names_html = []
for course_id in course_ids: 
  course = get_course_by_id(CourseLocator.from_string(course_id)) 
  course_names.append(course.display_name_with_default)
  course_names_html.append("<li>"+ str(course.display_name_with_default)+"</li>")
  # course_names_html.append("<li>"+ str(course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace'))+"</li>")

# WRITE XLS
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
# wb = Workbook(encoding='utf-8')
sheet = wb.active
sheet.title= 'Rapport'
filename = '/home/edxtma/csv/{}_Deeptech_for_business.xls'.format(timestr)
headers = ['ID', 'Prénom', 'Nom', 'Temps passé (min)']
first = True

j=2
for i, header in enumerate(headers):
  sheet.cell(1, i+1, header)

for key, user in all_users_data.items():
  i = 4 
  sheet.cell(j, 1, user['general']['username'])
  sheet.cell(j, 2, user['general']['firstname'])
  sheet.cell(j, 3, user['general']['lastname'])
  sheet.cell(j, 4, user['general']['time_tracking'])

  correctedExamGrade = 0
  sheet.cell(j, i+1, correctedExamGrade)
  first = False
  j += 1

sheet.cell(1, i+1, 'Note finale')

log.info('------------> Finish calculate grades and write xlsx report')

# SEND MAILS
output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ str(course_names_html) +"  <br/>Pour la période des 30 derniers jours uniquement.<br/><br/>Bonne r&eacute;ception<br>L'&eacute;quipe NETEXPLO<br></p></body></html>"

for email in emails:
  
  if not at_least_one_student :
    log.info('at_least_one_student :')
    log.info(at_least_one_student)
    break

  part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
  fromaddr = "ne-pas-repondre@themoocagency.com"
  msg = MIMEMultipart()
  msg['From'] = fromaddr
  msg['To'] = email
  msg['Subject'] = "NETEXPLO - " + str(course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace')) + ' - last 30 days - ' + str(time.strftime("%d.%m.%Y"))
  attachment = _files_values
  part = MIMEBase('application', 'octet-stream')
  part.set_payload(attachment)
  encoders.encode_base64(part)
  part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
  msg.attach(part)
  server = smtplib.SMTP('mail3.themoocagency.com', 25)
  server.starttls()
  server.login('contact', 'waSwv6Eqer89')
  msg.attach(part2)
  text = msg.as_string()
  server.sendmail(fromaddr, email, text)
  server.quit()
  log.info('Email sent to '+str(email))






# Command to execute: 
# source /edx/app/edxapp/edxapp_env && /edx/app/edxapp/edx-platform/manage.py lms shell < /edx/app/edxapp/edx-themes/deeptechforbusiness/lms/utils/script_new_users_deeptech.py

