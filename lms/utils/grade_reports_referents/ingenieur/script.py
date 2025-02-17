# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()
import datetime

import report_config
#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################


import json
import time
import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from lms.djangoapps.courseware.courses import get_course_by_id
# from lms.djangoapps.wul_apps.best_grade.helpers import check_best_grade
from lms.djangoapps.wul_apps.models import WulCourseEnrollment
from lms.djangoapps.courseware.models import StudentModule


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


def extract_field(data,fieldname):
  empty_str = "n/a"
  return json.loads(data).get(fieldname,empty_str)


import logging
log = logging.getLogger()


emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")

all_users_data = {}

umn_scorm_list = report_config.umn_scorm_list
email_count = 0

empty_str = "n.a."
empty_str = "n/a"

print("Collecting user data...")
timestamp_startup = round(time.time() * 1000)

for course_id in course_ids:
  print(course_id)
  print("course " + course_id, end="")
  course_key = CourseLocator.from_string(course_id)
  # course = get_course_by_id(course_key)
  course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)
  # course_name = course.display_name_with_default

  course_data = {}
  referents_list = set()
  for i in range(len(course_enrollments)):
  #for i in range(30):
    user = course_enrollments[i].user
    user_data = {}

    enrollment = course_enrollments[i]
    # skip staff
    if str(user.email).find('@yopmail') != -1 or str(user.email).find('@weuplearning') != -1 or str(user.email).find('@themoocagency') != -1 :
      continue
    #print(user.email)
    
    
    try:
      user_data["email"] = user.email
    except:
      user_data["email"] = extract_field(user.profile.custom_field, "email")

        
    try:
      user_data["date_joined"] = user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
    except:
      user_data["date_joined"] = empty_str

    try:
      user_data["last_login"] = user.last_login.strftime('%Y-%m-%d %H:%M:%S')
    except:
      user_data["last_login"] = empty_str



    user_data["first_name"] = extract_field(user.profile.custom_field, "first_name")
    user_data["last_name"] = extract_field(user.profile.custom_field, "last_name")


    user_data["structure"] = extract_field(user.profile.custom_field, "structure")


    user_data["status"] = extract_field(user.profile.custom_field, "status")
    user_data["preparedDiploma"] = extract_field(user.profile.custom_field, "preparedDiploma")


    user_data["school"] = extract_field(user.profile.custom_field, "school")
    user_data["formation"] = extract_field(user.profile.custom_field, "formation")
    user_data["class"] = extract_field(user.profile.custom_field, "class")
    user_data["year"] = extract_field(user.profile.custom_field, "year")
    user_data["referent"] = extract_field(user.profile.custom_field, "referent")
    user_data["update_marker"] = extract_field(user.profile.custom_field, "update_marker")
    user_data["timetracking"] = extract_field(user.profile.custom_field, "timetracking")
    user_data["diplomalvl"] = extract_field(user.profile.custom_field, "diplomalvl")
    user_data["schoolregion"] = extract_field(user.profile.custom_field, "schoolregion")
    user_data["regions"] = extract_field(user.profile.custom_field, "regions")

    ref = user_data["referent"]
    if '@' in ref:
      referents_list.add(ref.lower())
    else:
      # Time saving - collect only users with a referent
      continue

    try:
      wul_course_enrollment = WulCourseEnrollment.objects.get(course_enrollment_edx__user=user, course_enrollment_edx__course_id=course_key)
      global_time_tracking = wul_course_enrollment.global_time_tracking
    except:
      global_time_tracking = 0
      

    user_data['timetracking'] = datetime.timedelta(seconds=global_time_tracking)


    # Grade
    list_of_student_modules = StudentModule.objects.filter(course_id__exact=course_id, module_type__exact="scorm", student_id=user.id).order_by().values('module_state_key', 'state')
    grade_list = []


    final_grade_score = 0
    for module in umn_scorm_list[course_id] :
      total_module_score = 0

      for scorm_id in umn_scorm_list[course_id][module] :
        module_score = empty_str

        for scorm in list_of_student_modules:

          if str(scorm['module_state_key']) == str(scorm_id) :
            dictionarisation  = json.loads(scorm["state"])

            try :
              module_score = round(dictionarisation['lesson_score'] * 100, 2)
              total_module_score += module_score
            except:
              module_score = empty_str

        grade_list.append(module_score)

      module_average = round(total_module_score / len(umn_scorm_list[course_id][module]), 2)
      final_grade_score += module_average
      grade_list.append(module_average)

    final_grade = round(final_grade_score / len(umn_scorm_list[course_id]), 2)
    grade_list.append(final_grade)



    user_data['grade_list'] = grade_list 
    data = { "general": user_data }
    course_data[str(user.id)]= data

  all_users_data[course_id]= course_data
  print("[OK]")

# log.info('------------> Finish fetching user data and answers')
# log.info('------------> Begin Calculate grades and write xlsx report')
timestamp_calculate = round(time.time() * 1000)
# WRITE XLS

for ref in referents_list:
  print("Processing referent '" + ref + "'")
  assigned_students_counter = 0

  timestr = time.strftime("%Y_%m_%d")
  wb = Workbook()
  sheet = wb.active
  sheet.title= 'Rapport de notes'

  filename_base =  "/home/edxtma/csv/" + str(timestr) + "_"+ report_config.org_name +"_"+ report_config.group_name + ".xlsx"
  filename =  filename_base
  #print(filename)

  headers = report_config.headers
  for i, header in enumerate(headers):
    bg_color = "007DFF"
    if(i < 4):
      bg_color = "993300"
    sheet.cell(1, i+1, header)
    sheet.cell(1, i+1).fill = PatternFill("solid", fgColor=bg_color)
    sheet.cell(1, i+1).font = Font(b=False, color="FFFFFF")

  j=2

  for k, course_id in all_users_data.items():

    for key, user in course_id.items():
      user_ref = str(user['general']['referent'])
      user_ref = user_ref.lower()
      if user_ref != ref:
        continue
      assigned_students_counter +=1
      sheet.cell(j, 1, user['general']['regions'] )
      sheet.cell(j, 2, user['general']['structure'] )
      sheet.cell(j, 3, user['general']['preparedDiploma'] )
      sheet.cell(j, 4, user['general']['status'] )

      #sheet.cell(j, 5, user['general']['update_marker'] )

      sheet.cell(j, 5, user['general']['first_name'])
      sheet.cell(j, 6, user['general']['last_name'])
      sheet.cell(j, 7, user['general']['email'])

      date_joined = 'n/a'
      try:
        date_joined = user['general']['date_joined']
        date_joined = date_joined[:-12]

      except:
        date_joined = user['general']['date_joined']

      last_login = 'n/a'
      try:
        last_login = user['general']['last_login']
        last_login = last_login[:-12]
      except:
        last_login = user['general']['date_joined']

      sheet.cell(j, 8, date_joined)
      sheet.cell(j, 9, last_login)

      sheet.cell(j, 10, user['general']['schoolregion'] )
      sheet.cell(j, 11, user['general']['school'] )

      sheet.cell(j, 12, user['general']['formation'] )
      sheet.cell(j, 13, user['general']['class'] )
      sheet.cell(j, 14, user['general']['diplomalvl'] )

      sheet.cell(j, 15, user['general']['year'] )
      sheet.cell(j, 16, user['general']['referent'] )
      sheet.cell(j, 17, user['general']['timetracking'] )



      i=17
      save_grade = 0
      for grade in user['general']['grade_list'] :
        i += 1
        save_grade = grade

        if grade != empty_str :
          percent = str(grade) + '%'
        else :
          percent = grade

        sheet.cell(j, i, percent.replace('.',','))

      if float(save_grade) >= 70 : 
        sheet.cell(j, i).fill = PatternFill("solid", fgColor="21ad73")
        sheet.cell(j, i).font = Font(b=False, color="FFFFFF")
      else:
        sheet.cell(j, i).fill = PatternFill("solid", fgColor="ED4D39")
        sheet.cell(j, i).font = Font(b=False, color="FFFFFF")

      j += 1


  course_names_html = []
  for course_id in course_ids: 
    course = get_course_by_id(CourseLocator.from_string(course_id)) 
    course_names_html.append("<li>"+ str(course.display_name_with_default)+"</li>")

  output = BytesIO()
  wb.save(output)
  _files_values = output.getvalue()
  course_names_html = ''.join(course_names_html)

  html = report_config.email_body
  print(str(assigned_students_counter) + " matching students")
  if(assigned_students_counter == 0):
    print("WARNING : " + ref + " NO STUDENTS")
  part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
  fromaddr = "ne-pas-repondre@themoocagency.com"
  
  msg = MIMEMultipart()
  msg['From'] = fromaddr
  
  split_ref = ref.split(";")
  for target_ref in split_ref:
    email_count +=1
    msg['To'] = target_ref
    print("Sending email to " + target_ref )

    msg['Subject'] = "Suivi du e-learning Passeport Nucléaire parcours Master/Ingénieur"
    #msg['Subject'] = "[DEBUG]Suivi du e-learning Passeport Nucléaire parcours Master/Ingénieur"
    attachment = _files_values
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
    msg.attach(part)
    server = smtplib.SMTP(report_config.mailer_addr, 25)
    server.starttls()
    server.login(report_config.mailer_login, report_config.mailer_password)
    msg.attach(part2)
    text = msg.as_string()
    #email = 'theo.gicquel@weuplearning.com'
    email = target_ref
    server.sendmail(fromaddr, email, text)
    #server.sendmail(fromaddr, 'theo.gicquel@weuplearning.com', text)
    #server.sendmail(fromaddr, 'melanie.zunino@weuplearning.com', text)

    server.quit()
    #log.info('Email sent to '+str(email))


#log.info('------------> Finish calculate grades and write xlsx report')
timestamp_exit = round(time.time() * 1000)
print("Complete report took  " + str(timestamp_exit - timestamp_startup) + " ms / " + str((timestamp_exit - timestamp_startup)/1000) + " sec : "   ) 
print("  - Compiling LMS data took " + str(timestamp_calculate - timestamp_startup) + " ms / " + str((timestamp_calculate - timestamp_startup)/1000) + " sec"   ) 
print("  - Writing XLS took " + str(timestamp_exit - timestamp_calculate) + " ms / " + str((timestamp_exit - timestamp_calculate)/1000) + " sec"   ) 
print("sent " + str(email_count) + " emails")

# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/umn/lms/utils/grade_reports_referents/ingenieur/script.py 'theo.gicquel@weuplearning.com' 'course-v1:umn+pi+01'
