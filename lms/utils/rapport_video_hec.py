# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import zipfile

import sys
importlib.reload(sys)

import os
from io import BytesIO

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################



import time
import json
import datetime

from opaque_keys.edx.locations import SlashSeparatedCourseKey
from opaque_keys.edx import locator
from lms.djangoapps.wul_apps.models import WulCourseEnrollment
from common.djangoapps.student.models import User, UserProfile
from lms.djangoapps.courseware.models import StudentModule
from student.models import CourseEnrollment
from opaque_keys.edx.keys import CourseKey
from courseware.courses import get_course_by_id

from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import logging
log = logging.getLogger()

emails = sys.argv[1].split(";")
#course_ids = sys.argv[2].split(";")
#course-v1:hec-pole-emploi+01+2022
course_ids = ["course-v1:hec-pole-emploi+IP_NEG+2023", "course-v1:hec-pole-emploi+IP+2023", "course-v1:hec-pole-emploi+NEG+2023", "course-v1:hec-pole-emploi+webinaire+2023"]
dict_course_name = {"course-v1:hec-pole-emploi+NEG+2023": "Negociation (NEG)", "course-v1:hec-pole-emploi+IP+2023": "Initiative Personnelle (IP)", "course-v1:hec-pole-emploi+IP_NEG+2023": "NEG+IP", "course-v1:hec-pole-emploi+webinaire+2023": "Webinaire"}

scorm_page_ids = {'course-v1:hec-pole-emploi+IP+2023': {'block-v1:hec-pole-emploi+IP+2023+type@scorm+block@9608f74c73564e9f855dc290ee0d8328': 'Scorm module'}, 'course-v1:hec-pole-emploi+IP_NEG+2023': {'block-v1:hec-pole-emploi+IP_NEG+2023+type@scorm+block@5c719e3bb8d048f2ae7dde0c177916c5': "Vidéo d'introduction", 'block-v1:hec-pole-emploi+IP_NEG+2023+type@scorm+block@dbb48f175f9c48f6baf8c2f35059b24a': 'Étude de cas', 'block-v1:hec-pole-emploi+IP_NEG+2023+type@scorm+block@790d688c96bd4e4393c5a4acb9b00408': 'Étude de cas 2', 'block-v1:hec-pole-emploi+IP_NEG+2023+type@scorm+block@af00ac1520ee41698dcc4f839e0656b5': 'Étude de cas 3'}, 'course-v1:hec-pole-emploi+NEG+2023': {'block-v1:hec-pole-emploi+NEG+2023+type@scorm+block@75d793663c1d4256aedd2ae063d3c978': 'Scorm module'}, 'course-v1:hec-pole-emploi+webinaire+2023': {}}


#dict_hardcoded_chapter_key = {"course-v1:hec-pole-emploi+01+2022": ['block-v1:hec-pole-emploi+01+2022+type@chapter+block@54f8633948664c0bb85f54b33947eaba', 'block-v1:hec-pole-emploi+01+2022+type@chapter+block@d9d5984d68d04f008df511ace8dc843f', 'block-v1:hec-pole-emploi+01+2022+type@chapter+block@5089741f072547ef97395ef7a5b0ee1c']}
dict_hardcoded_chapter_key = {'course-v1:hec-pole-emploi+IP+2023': ['block-v1:hec-pole-emploi+IP+2023+type@chapter+block@210326b6198446868cda2d4e2028e291', 'block-v1:hec-pole-emploi+IP+2023+type@chapter+block@43078e3916414a46b72fe6321e1f3d25', 'block-v1:hec-pole-emploi+IP+2023+type@chapter+block@aa33f1dae86643dcb2b6d9ccb6f81d17', 'block-v1:hec-pole-emploi+IP+2023+type@chapter+block@111b7f2672ff42c585dcdee7e9b4bca6'], 'course-v1:hec-pole-emploi+IP_NEG+2023': ['block-v1:hec-pole-emploi+IP_NEG+2023+type@chapter+block@5719ed51d13842efa6ff46564a3a2430', 'block-v1:hec-pole-emploi+IP_NEG+2023+type@chapter+block@c1e54fc31ee94f878ea6bd9959a5f15b', 'block-v1:hec-pole-emploi+IP_NEG+2023+type@chapter+block@9d9193c1b488439a991aadc2799955bb'], 'course-v1:hec-pole-emploi+NEG+2023': ['block-v1:hec-pole-emploi+NEG+2023+type@chapter+block@6ccc988be0b84e7b82f4e92bdb35101b'], 'course-v1:hec-pole-emploi+webinaire+2023': []}

all_courses_video_student_module = []

users_data = dict()

users_per_course = dict()

list_chapters_name = dict()

import string
from opaque_keys.edx.keys import CourseKey, UsageKey
from opaque_keys.edx.locator import BlockUsageLocator
from common.lib.xmodule.xmodule.modulestore.django import modulestore

# entry point to the block_structure api.
from openedx.core.djangoapps.content.block_structure.api import get_course_in_cache

all_user_set = set()
list_of_student_scorms = list()
for course_id in course_ids :
    list_of_student_modules = StudentModule.objects.filter(course_id__exact=course_id).order_by().values('student_id', 'module_state_key', 'state')
    list_of_student_scorms = StudentModule.objects.filter(course_id__exact=course_id, module_type="scorm")
    #log.info(f"problem of course :{course_id} => {list_of_student_problems}")
    users = set()
    users_scorms = set()
    for student_module in list_of_student_modules:
        all_courses_video_student_module.append(student_module)
        users.add(User.objects.get(id = student_module["student_id"]))
    all_user_set.update(users)
    users_per_course[course_id] = users
    log.info(course_id)


videos_list = list()

def convert_str_to_obj(saved_video_position):
    str_time_video = saved_video_position.replace('"saved_video_position": "', '').replace('{', '').replace('}', '').replace('"', '')
    if 'speed' in str_time_video:
        char_to_replace = ["0.25", "0.5", "0.75","1.0","1.25","1.5","1.75","2.0", ",", " ","speed:"]
        for char in char_to_replace:
            str_time_video = str_time_video.replace(char, "")
    return str_time_video

def convert_str_to_int(time_str):
    """Get seconds from time."""
    h, m, s = time_str.split(':')
    return int(h) * 3600 + int(m) * 60 + int(s)

for course_id, users in users_per_course.items():
    user_data = dict()
    users_data = dict()
    users_scorm_completion = dict()
    for index, user in enumerate(users):
        
    #uncomment this lines for testing, 
        # log.info(index)
        # if index == 100:
        #     break
        #log.info(index)
        #if not user.email.find("annojan.kandiah@weuplearning.com")!= -1:
        #    continue
        
        #course_enrollment = CourseEnrollment.objects.get(course_id=CourseKey.from_string(course_ids[0]), user=user)
        #course_key = SlashSeparatedCourseKey.from_string(str(course_id))
        #course = get_course_by_id(course_key)
        # Escape fake email address
        if user.email.find("@example")!= -1 or user.email.find("@themoocagency") != -1 or user.email.find("@weuplearning")!= -1 or user.email.find("@yopmail")!= -1 or user.email.find("@amazon")!= -1 or user.email.find("@fake")!= -1:
            continue

        #log.info(dir(user))
        user_row = dict()
        try:
            user_data["id"] = user.id
        except:
            user_data["id"] = ""

        try:
            user_data["username"] = user.username
        except:
            user_data["username"] = ""

        try:
            user_data["email"] = user.email
        except:
            user_data["email"] = ""

        try:
            user_data["date_joined"] = user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
        except:
            user_data["date_joined"] = ""

        try:
            user_data["last_login"] = user.last_login.strftime('%Y-%m-%d %H:%M:%S')
        except:
            user_data["last_login"] = ""

        try:
            user_data["name"] = user.name
        except:
            user_data["name"] = ""

        try:
            user_data["first_name"] = user.first_name
        except:
            user_data["first_name"] = ""

        try:
            user_data["end_date_enrollment"] = (user.date_joined + datetime.timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        except:
            user_data["end_date_enrollment"] = ""

        user_row = []
        video_dict = dict()
        user_data["total_video_time"] = datetime.timedelta(seconds=0)

        log.info('treating user :')
        log.info(user.email)

        #try:
        course_key = locator.CourseLocator.from_string(str(course_id))
        collected_block_structure = get_course_in_cache(course_key)
        
        user_scorms = StudentModule.objects.filter(student=user, course_id__exact=course_id, module_type="scorm")
        #log.info(dir(user_scorm))
        for scorm_id, scorm_name in scorm_page_ids[course_id].items():
            users_scorm_completion[scorm_id] = "no"
            for scorm in user_scorms:
                log.info(scorm.done)
                if scorm_id == scorm.module_state_key.block_id:
                    users_scorm_completion[scorm_id] = scorm.done
        #except:
        #    continue

        try:
            course_key = locator.CourseLocator.from_string(str(course_id))
            collected_block_structure = get_course_in_cache(course_key)
            list_chapter_course = list()
            for chapter_key in dict_hardcoded_chapter_key[course_id]:
                usagekey = UsageKey.from_string(chapter_key)
                chapter_name = collected_block_structure.get_xblock_field(usagekey, "display_name")
                detailed_time_tracking = json.loads(WulCourseEnrollment.get_enrollment(user=user, course_id=course_id).detailed_time_tracking)
                chapter_key = chapter_key.split("@")[2]
                if chapter_key in detailed_time_tracking:
                    user_data[chapter_name] = datetime.timedelta(seconds=detailed_time_tracking[chapter_key])
                else:
                    user_data[chapter_name] = datetime.timedelta(seconds=0)
                list_chapter_course.append(chapter_name)
            list_chapters_name[course_id] = list_chapter_course
        except:
            continue

        # Access TimeTracking for every courses
        global_time_tracking_cumul = 0

        try:
            detailed_time_tracking = json.loads(WulCourseEnrollment.get_enrollment(user=user, course_id=course_id).detailed_time_tracking)
            #keys = set(scorm_page_ids).intersection(detailed_time_tracking)
            #user_detailed_time_tracking = {k:detailed_time_tracking[k] for k in keys}
            user_detailed_time_tracking = {key: detailed_time_tracking[key] for key in scorm_page_ids if key in detailed_time_tracking}
            if user_detailed_time_tracking:
                total_time = user_detailed_time_tracking.values()
                user_data["scorm_time_tracking"] = datetime.timedelta(seconds=sum(total_time))
            else:
                user_data["scorm_time_tracking"] = datetime.timedelta(seconds=0)
        except:
            user_data["scorm_time_tracking"] = datetime.timedelta(seconds=0)

        try:
            course_key = locator.CourseLocator.from_string(str(course_id))
            wul_course_enrollment = WulCourseEnrollment.objects.get(course_enrollment_edx__user=user, course_enrollment_edx__course_id=course_key)
            global_time_tracking = wul_course_enrollment.global_time_tracking
            global_time_tracking_cumul += global_time_tracking

        except:
            pass


        if global_time_tracking_cumul == 0 :
            user_data["global_time_tracking"] = datetime.timedelta(seconds=0)
        else:
            user_data["global_time_tracking"] = datetime.timedelta(seconds=global_time_tracking_cumul)

        # Get detailled video information from student modules
        for video in videos_list :
            video_dict[video] = "Non"

        total_video_seconds = 0
        try:
            for result in all_courses_video_student_module:
                if(user.id == result["student_id"]):
                    if str(result["module_state_key"]).split("+")[4] in video_dict:
                        user_state = result["state"]
                        video_time_tracking = convert_str_to_obj(user_state)

                        time_video_time_tracking = convert_str_to_int(video_time_tracking)
                        total_video_seconds += time_video_time_tracking
                        
                        video_dict[str(result["module_state_key"]).split("+")[4]] = datetime.timedelta(seconds=time_video_time_tracking)
                        time_video_time_tracking += time_video_time_tracking
                        user_data["total_video_time"] = datetime.timedelta(seconds=time_video_time_tracking)
                    else:
                        print('ERROR {} VIDEO NOT IN THE LIST !'.format(video_dict[str(result["module_state_key"]).split("+")[4]]))
        except Exception as e:
            print(str(user.id)+" : error with watch a video field")
            print(e)

        try:
            # TODO retrive grade of current user
            user_data["grade"] = "N/A"
        except:
            user_data["grade"] = "N/A"

        if course_id == "course-v1:hec-pole-emploi+IP_NEG+2023":
            user_row = [user_data["username"],user_data["email"],user_data["first_name"],user_data["name"],user_data["date_joined"],user_data["end_date_enrollment"],user_data["last_login"], user_data["global_time_tracking"], "N/A", user_data["total_video_time"]]
        else:
            user_row = [user_data["username"],user_data["email"],user_data["first_name"],user_data["name"],user_data["date_joined"],user_data["end_date_enrollment"],user_data["last_login"], user_data["global_time_tracking"], "N/A", user_data["total_video_time"], user_data["grade"]]

        for scorm_id in users_scorm_completion:
            log.info(users_scorm_completion[scorm_id])
            user_row.append(users_scorm_completion[scorm_id])
        if course_id == "course-v1:hec-pole-emploi+IP_NEG+2023":
            user_row.extend(["N/A", "N/A", "N/A"])
        for chapter_name in list_chapter_course:
            user_row.append(user_data[chapter_name])
        #*************************************************************************************
        # TODO IMPORTANT ! put this in custom field then remove it after july 19
        #*************************************************************************************
        # july_19_date = datetime.datetime(2022, 7, 19)
        # new_user = july_19_date.timestamp() < user.date_joined.timestamp()

        # calculated_time = total_video_seconds*10
        # total_video_time =datetime.timedelta(seconds=calculated_time)
        # user_data["calculated_time"] = total_video_time

        # custom_field['total_time_calculated'] = calculated_time
        # user.profile.custom_field = json.dumps(custom_field)
        # user.profile.save()


        # if new_user:
        #     user_data["calculated_time"] = datetime.timedelta(0,0,0)
        # else :
        #     if custom_field and 'total_time_calculated' in custom_field.keys():
        #         user_data["calculated_time"] = datetime.timedelta(seconds=custom_field['total_time_calculated'])
        #     else:
        #         user_data["calculated_time"] = datetime.timedelta(0,0,0)
        # if user_data["global_time_tracking"] != 'n/a':
        #     user_data["global_and_calculated_time"] = total_video_time + user_data['global_time_tracking']
        # else:
        #     user_data["global_and_calculated_time"] = total_video_time

        # user_row.append(user_data["calculated_time"])
        # user_row.append(user_data["global_and_calculated_time"])

        users_data[user.username.capitalize()] = user_row
    users_per_course[course_id] = users_data

#
## Print excel file
row = 1


## Workbook
wb = Workbook()
wb.remove(wb.active)

def create_sheet(sheet_name, users, workbook, course_id):
    common_header = ["Username","Email","Prénom","Nom","Date de création de compte","Date de fin d'accès","Date de dernière connexion", "Temps passé total", "Webinaire finalisé", "Temps passé webinaire", "Note obtenue"] 

    for scorm_id, scorm_name in scorm_page_ids[course_id].items():
        log.info(scorm_name)
        common_header.append(scorm_name + " validé")
    if course_id == "course-v1:hec-pole-emploi+IP_NEG+2023":
        log.info(f"default header: {common_header}")
        del common_header[10]
        log.info(f"poped header: {common_header}")
        common_header += ["Note NEG", "Note IP", "Score Moyen"]
        log.info(f"modifed header:{common_header}")
        #common_header.append(specific_neg_ip_header)
    sheet = workbook.create_sheet(sheet_name)
    users_of_a_course = users_per_course[course_id]
    if not dict_hardcoded_chapter_key[course_id]:
        return
    for chapter_name in list_chapters_name[course_id]:
        common_header.append(chapter_name)
    for i, header in enumerate(common_header):
        sheet.cell(row=1, column=(i+1)).value = header
    j=2
    for user in users:
        user_row = user[1]
        l=0
        for value in user_row :
            sheet.cell(row=j, column=(l+1)).value = value
            l=l+1
        j=j+1

for course_id in course_ids:
    users_data = users_per_course[course_id]
    if users_data == []:
        continue
    ordered_users = sorted(users_data.items(), key=lambda x: x[1])
    create_sheet(dict_course_name[course_id], ordered_users, wb, course_id)

timestr = time.strftime("%Y_%m_%d")
filename = "Rapport_hec_{}.xlsx".format(timestr)
filepath = '/edx/var/edxapp/media/microsites/hec-pole-emploi/reports/{}'.format(filename)
wb.save(filepath)
output = BytesIO()
_files_values = output.getvalue()
html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de donn&eacute;es Amazon"

## Send email
for email in emails:
    part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "HEC pole-emploi <ne-pas-repondre@themoocagency.com>"
    toaddr = email
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "Rapport temps passé HEC pole-emploi"

    attachment = _files_values

    with open(filepath, 'rb') as f:
        attachment = f.read()
        
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= {}".format(filename))
    msg.attach(part)

    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()

    print('Email sent to ',toaddr)


## delete old files
two_weeks_ago = datetime.datetime.today() - datetime.timedelta(days=14)
try:
    os.remove('/edx/var/edxapp/media/microsites/hec-pole-emploi/reports/Rapport_hec_{}.xlsx'.format(two_weeks_ago.strftime("%Y_%m_%d")))
except:
    pass

#/edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/hec-pole-emploi/lms/utils/rapport_video_hec.py "annojan.kandiah@weuplearning.com"
#/edx/app/edxapp/venvs/edxapp/bin/python /home/ubuntu/hec_scripts/rapport_video_hec.py "annojan.kandiah@weuplearning.com"
