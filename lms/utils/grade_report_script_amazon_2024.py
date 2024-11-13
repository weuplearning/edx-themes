
# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import zipfile

import sys
importlib.reload(sys)

import os
from io import BytesIO

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################



import time
import datetime
import json

# from opaque_keys.edx.locations import SlashSeparatedCourseKey
from opaque_keys.edx import locator
from lms.djangoapps.wul_apps.models import WulCourseEnrollment
from opaque_keys.edx.locator import CourseLocator
from lms.djangoapps.courseware.courses import get_course_by_id

from lms.djangoapps.wul_apps.best_grade.helpers import check_best_grade

from common.djangoapps.student.models import User, UserProfile
# from lms.djangoapps.courseware.models import StudentModule
from student.models import CourseEnrollment


from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import logging
log = logging.getLogger()



emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")

## Workbook
wb = Workbook()
ws = wb.active
ws.title = "Grading report"


## Construct data
users = User.objects.all()
users_data = dict()
siret = dict()

# Headers
headers = ["Nome utente","Email","Data di creazione", "Data dell'ultima connessione","Tempo trascorso","Numero di corsi in corso","Numero di corsi completati","Export","Logistica e sostenibilità","Introduzione all'e-commerce e alla vendita su Amazon","Costruire la presenza del tuo brand su Amazon","Principi di gestione per il tuo business","Avvia la tua attività","Fondamenti di marketing","Strategie di Marketing","Primi passi nella vendita digitale","Vendite online internazionali","Le basi della creazione del marchio","Le basi del content marketing","Introduzione alla Finanza aziendale","Gestire la finanza aziendale","Elementi di finanza aziendale","Comprendere le metriche aziendali e l'importanza del cliente","Come analizzare la propria base clienti","Creare la tua startup e vendere online","Creare un piano di sviluppo per la tua attività","Panoramica delle operazioni commerciali, logistica e metodi di spedizione","Strategia di vendita digitale","Definire la tua strategia commerciale","Creare valore per la tua attività","Come assumere nuovi talenti","Stile di Leadership e di Management"]



today = datetime.datetime.now(datetime.timezone.utc)



### Loop over all_user 


for index, user in enumerate(users):

    ## uncomment this lines for testing, 
    # if index == 250:
    #    break


    # Escape fake email address
    if user.email.find("@example")!= -1 or user.email.find("@themoocagency") != -1 or user.email.find("@weuplearning")!= -1 or user.email.find("@yopmail")!= -1 or user.email.find("@amazon")!= -1 or user.email.find("@fake")!= -1:
        continue



    user_data = dict()

    user_data["name"] = user.profile.name
    try:
        user_data["id"] = user.id
    except:
        user_data["id"] = ""
    try:
        user_data["username"] = user.username
    except:
        user_data["username"] = ""
    try:
        user_data["email"] = user.email
    except:
        user_data["email"] = ""

    custom_field = json.loads(user.profile.custom_field)
    try:
        user_data["date_joined"] = user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
    except:
        user_data["date_joined"] = ""
    try:
        user_data["last_login"] = user.last_login.strftime('%Y-%m-%d %H:%M:%S')
    except:
        user_data["last_login"] = ""


    user_row = []
    video_dict = dict()
    user_data["enrolled_to"] = 0
    user_data["finished_course"] = 0
    user_data["total_video_views"] = 0


    ### Grade Data

    global_time_tracking_cumul = 0


    for course_id in course_ids :

        all_course_enrollment = CourseEnrollment.objects.filter(user=user)
        user_data[course_id] = ''


        for enrollment in all_course_enrollment :


            if str(course_id) == str(enrollment.course_id) :

                course_key = CourseLocator.from_string(course_id)
                course = get_course_by_id(course_key)

                user_data["enrolled_to"] += 1

                #log.info(course_id)
                try:
                    gradesTest = check_best_grade(user, course, force_best_grade=True)
                    user_data[course_id] = gradesTest.summary['percent']

                    if gradesTest.summary['percent'] >= 0.7 :
                        user_data["finished_course"] += 1
                except:
                    user_data[course_id] = 'Pas noté'


                try:
                    course_key = locator.CourseLocator.from_string(str(course_id))
                    wul_course_enrollment = WulCourseEnrollment.objects.get(course_enrollment_edx__user=user, course_enrollment_edx__course_id=course_key)

                    global_time_tracking = wul_course_enrollment.global_time_tracking
                    global_time_tracking_cumul += global_time_tracking
                except:
                    pass


    ### TimeTracking Data

    if global_time_tracking_cumul == 0 :
        user_data["global_time_tracking"] = 'n/a'
    else:
        user_data["global_time_tracking"] = datetime.timedelta(seconds=global_time_tracking_cumul)




    user_row = [user_data["username"],user_data["email"],user_data["date_joined"],user_data["last_login"],user_data["global_time_tracking"],user_data["enrolled_to"],user_data["finished_course"]]
    # user_row = [user_data["username"],user_data["email"],user_data["name"],user_data["region"],user_data["siret"],user_data["phone_number"],user_data['online_sales'],user_data["date_joined"],user_data["last_login"],user_data["global_time_tracking"],user_data["enrolled_to"],user_data["finished_course"]]

    for course_id in course_ids :
        user_row.append(user_data[course_id])


    users_data[user.username.capitalize()] = user_row

ordered_users = sorted(users_data.items(), key=lambda x: x[1])


### Print excel file


row = 1

sheet = wb.active
for i, header in enumerate(headers):
    sheet.cell(1, (i+1), header)
    sheet.cell(1, i+1).fill = PatternFill("solid", fgColor="1E2631")
    sheet.cell(1, i+1).font = Font(b=True, color="BA4926")
j=2

for user in ordered_users:
    user_row = user[1]
    l=0
    for value in user_row :
        sheet.cell(row=j, column=(l+1)).value = value
        l=l+1
    j=j+1

timestr = time.strftime("%Y_%m_%d")
filename = "Amazon_ratings_report_{}.xlsx".format(timestr)
filepath = '/home/ubuntu/amazon_reports/{}'.format(filename)
wb.save(filepath)


### Create a new zip file and write the Excel file into it


zipname = "ratings_report.zip"
zippath = '/home/ubuntu/amazon_reports/{}'.format(zipname)

with zipfile.ZipFile(zippath, 'w', zipfile.ZIP_DEFLATED, compresslevel=9) as myzip:
    myzip.write(filepath, arcname=filename)

output = BytesIO()
wb.save(output)
_files_values = output.getvalue()
html = "<html><head></head><body><p>Hello,<br/><br/>Please find attached the Accelera con Amazon data reports.<br/><br/>Kind regards,<br />The WeUp Learning team</html>"


### Send email


for email in emails:

    part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "Amazon <ne-pas-repondre@themoocagency.com>"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = email
    msg['Subject'] = "Accelera con Amazon Data Reports"

    attachment = _files_values

    # Load your zip file instead of the Excel file
    with open(zippath, 'rb') as f:
        attachment = f.read()

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= {}".format(zipname))
    msg.attach(part)

    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, email, text)
    server.quit()

    print('Email sent to ',email)


## delete old files
two_weeks_ago = datetime.datetime.today() - datetime.timedelta(days=14)
try:
    os.remove('/home/ubuntu/amazon_reports/Amazon_ratings_report_{}.xlsx'.format(two_weeks_ago.strftime("%Y_%m_%d")))
except:
    pass



#qualif
# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/amazon-italie/lms/utils/grade_report_script_amazon_2024.py "cyril.adolf@weuplearning.com" "course-v1:amazon-it+definire_strategia_commercialeAvanzamento+az_22"

# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/amazon-italie/lms/utils/grade_report_script_amazon_2024.py "fpelli@amazon.it;melanie.zunino@weuplearning.com" "course-v1:amazon-it+export+az_01;course-v1:amazon-it+logistica_e_sostenibilita+az_02;course-v1:amazon-it+introduzione_e-commerce_vendita+az_03;course-v1:amazon-it+presenza_brand+az_04;course-v1:amazon-it+principi_di_gestione+az_05;course-v1:amazon-it+avvia_la_tua_attivita+az_06;course-v1:amazon-it+fondamenti_di_marketing+az_07;course-v1:amazon-it+strategie_di_marketing+az_08;course-v1:amazon-it+primi_passi_vendita_digitale+az_09;course-v1:amazon-it+vendite_online_internazionali+az_10;course-v1:amazon-it+basi_creazione_marchio+az_11;course-v1:amazon-it+basi_content_marketing+az_12;course-v1:amazon-it+introduzione_finanza_aziendale+az_13;course-v1:amazon-it+gestire_finanza_aziendale+az_14;course-v1:amazon-it+elementi_finanza_aziendale+az_15;course-v1:amazon-it+metriche_aziendali_importanza_cliente+az_16;course-v1:amazon-it+analizzare_propria_base_clienti+az_17;course-v1:amazon-it+creare_startup_vendere_online+az_18;course-v1:amazon-it+piano_sviluppo_attivita+az_19;course-v1:amazon-it+panoramica_operazioni_logistica_spedizione+az_20;course-v1:amazon-it+strategia_vendita_digitale+az_21;course-v1:amazon-it+definire_strategia_commerciale+az_22;course-v1:amazon-it+creare_valore_attivita+az_23;course-v1:amazon-it+assumere_nuovi_talenti+az_24;course-v1:amazon-it+stile_leadership_management+az_25"

