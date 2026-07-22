# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################


import json
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from lms.djangoapps.courseware.courses import get_course_by_id
from lms.djangoapps.wul_apps.best_grade.helpers import check_best_grade
from lms.djangoapps.wul_apps.models import WulCourseEnrollment

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import logging
log = logging.getLogger()


emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")


buggued_tt = []

all_users_data = {}
headers = ['Email', 'Nom complet', 'Adresse', 'Code postal', 'Ville',  'Région', 'Profession', 'Profession si autre', 'Jour de connexion et temps', 'Temps passé par module', 'Progression']

for course_id in course_ids:
    log.info('------------> Start calculate grades for course : '+str(course_id))

    course_key = CourseLocator.from_string(course_id)
    course = get_course_by_id(course_key)
    course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)

    for i in range(len(course_enrollments)):
        user = course_enrollments[i].user
        enrollment = course_enrollments[i]

        if str(user.email).find('@yopmail') != -1 or str(user.email).find('@weuplearning') != -1 or str(user.email).find('@themoocagency') != -1 :
            continue

        # TimeTracking
        try:
            wul_course_enrollment = WulCourseEnrollment.objects.get(course_enrollment_edx__user=user, course_enrollment_edx__course_id=course_key)
            daily_time_tracking = json.loads(wul_course_enrollment.daily_time_tracking)
            detailled_time_tracking = json.loads(wul_course_enrollment.detailed_time_tracking)
        except:
            daily_time_tracking = {}
            detailled_time_tracking = {}


        user_id = str(user.id)

        if user_id not in all_users_data:
            not_found_str = 'n.a.'
            user_data = {}
            user_data["email"] = user.email
            user_data["name"] = user.profile.name
            user_data["adress"] = json.loads(user.profile.custom_field).get('adress', not_found_str)
            user_data["post_code"] = json.loads(user.profile.custom_field).get('post_code', not_found_str)
            user_data["city"] = json.loads(user.profile.custom_field).get('city', not_found_str)
            user_data["region"] = json.loads(user.profile.custom_field).get('region', not_found_str)
            user_data["profession"] = json.loads(user.profile.custom_field).get('profession', not_found_str)
            user_data["profession_autre"] = json.loads(user.profile.custom_field).get('profession_autre', not_found_str)

            all_users_data[user_id] = {
                "profil": user_data,
                "courses": {}
            }

        # Grade (spécifique au cours)
        user_grade = {}
        gradesTest = check_best_grade(user, course, force_best_grade=True)
        userPersentGrade = gradesTest.summary['percent']

        try:
            user_grade['global'] = round(userPersentGrade * 100, 2)
        except:
            user_grade['global'] = 0

        all_users_data[user_id]["courses"][course_id] = {
            "grades": user_grade,
            "tt_detailled": detailled_time_tracking,
            "tt_daily": daily_time_tracking
        }



correspondance_section_tt = {
    # # occitanie
    # "e73e91fe60fc450789f0a5faf244143a" : "Introduction",
    # "1939587bf96b484186bb524099cad8f5" : "Etape 1 : Dépistage",
    # "d53973250d4f463d99e10bcaf7f0a95c" : "Etape 2 : Gestion des alertes",
    # "4a22fe0f40e94dfbaf6edd6d1250261c" : "Etape 2 : Evaluation approfondie",
    # "ae38e833526c4779909f40ca4e8d24eb" : "Cas clinique",
    # "8ee6bd2d4c2a488fab130ad36fe67156" : "Présentation de l'outil de coordination régional",
    # "a6bdc40c399b4c66ad869d216b0dc7ce" : "Conclusion",

    # Etape 1 : Dépistage
    'course-v1:icope+e1+2026': {

        "4f2d1a15e4b74d259a958ffc1417a279" : "Introduction depistage",
        "0737e83eb32d40bbb70e4efe874f1cfe" : "Dépistage",
        "49066ca56fb44f0c8fac4773017b5c38" : "Conclusion depistage",
    },
    # Etape 2 : Gestion des alertes
    'course-v1:icope+gda+2026' : {
        "57fa9317431f475d856f4c556d633ff8" : "Introduction gestion des alertes",
        "604c1d346e674204ab704d741e7bebd5" : "Gestion des alertes",
        "f64f3a4f6c6b4d12a5584afb55e554d3" : "Conclusion gestion des alertes",
    },
    # Etape 2 : Evaluation approfondie
    'course-v1:icope+ea+2026' : {
        "cbab83172a1c4dd6adf5a4c44dd68d33" : "Introduction evaluation approfondie",
        "ced24a37e0c74be1958406c745eb6267" : "Évaluation approfondie",
        "9197235e1e22459d8002f7cb65d9adb1" : "Conclusion evaluation approfondie",
    },
    # Soins bucco-dentaires
    'course-v1:icope+soins_bucco_dentaires+2024' : {
        "4ec7d84360ba489e80925e0c59f7ef92" : "Formation bucco-dentaires"
    }
}
if 'course-v1:audition-icope+med+01' in course_ids : 
    correspondance_section_tt = {
        "f7f3e683f4d74b3f8baa560e5c71f1e8" : "Physiopathologie de l'oreille",
        "a5981a4c949347c98fe169cb88f59134" : "Du dépistage de la surdité à la décision",
        "1c97afb70dce4ebaa1c5848df2daae98" : "Environnement des mesures auditives",
        "2a8fa472c3974e64ad940aff67abcf58" : "Acoumétrie",
        "01f282202fd547c696adaff0b07d6d08" : "Audiométrie tonale avec assourdissement",
        "f20382e487fd49ff9cf3d1ade5ddb132" : "Audiométrie vocale dans le silence et corrélation tonale/vocale",
        "96f431656590437a8e6981d09370e4d3" : "Audiométrie vocale dans le Bruit",
        "f4e42a79c12f47ba83e287d226474d42" : "Formation à l'otoscopie et à la gestion des difficultés",
        "d13557ba965f4ad398831929ba361960" : "Orientation thérapeutique",
        "e832ba93bd90442c8f280e0f4bbd6240" : "Orientation vers la gériatrie",
        "0c61635589cf4f439babe61c7b62de45" : "Serious Game",
        "8d5a1a2021cb47f89850178aa0849f35" : "Evaluation finale",
    }


# WRITE EXCEL AND SEND MAILS
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= 'Rapport de notes'
filename = '/home/edxtma/csv/Icope_grade_report_{}.xlsx'.format(timestr)


j=0
for user_id, user_entry in all_users_data.items():
    i=0
    j+=1
    profil = user_entry['profil']

    for i, header in enumerate(headers):
        sheet.cell(j, i+1, header)
        sheet.cell(j, i+1).fill = PatternFill("solid", fgColor="6B9AAF")
        sheet.cell(j, i+1).font = Font(b=False, color="FFFFFF")

    j+=1
    i=0
    for key, value in profil.items():
        sheet.cell(j, i+1, value)
        i+=1

    for course_id, course_data in user_entry['courses'].items():

        percent_global = str(course_data['grades']['global']) + '%'
        sheet.cell(j, i+3, percent_global)
        spacer = 0
        
        for id, section_name in correspondance_section_tt[course_id].items():
            empty=True
            for hash, seconds in course_data['tt_detailled'].items():

                if hash == id:
                    sheet.cell(j, i+2, str(section_name) + " - " + str(round(seconds/60))+" min")
                    empty=False
                    spacer += 1
                    j+=1
                    break
    
            if empty :
                sheet.cell(j, i+2, str(section_name) + " - 0 min")
                spacer += 1
                j+=1

        j -= spacer

        for day, seconds in course_data['tt_daily'].items():
            if seconds > 20000 :
                buggued_tt.append(user_entry['profil']['email'])

            sheet.cell(j, i+1, str(day) + " : " + str(round(seconds/60))+" min")
            j+=1

        if len(course_data['tt_daily'].items()) >= spacer:
            j+= 1
        else:
            j+= (spacer+1) - len(course_data['tt_daily'].items())




output = BytesIO()
wb.save(output)
_files_values = output.getvalue()

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de temps passé pour les différents cours de la plateforme icope-formation <br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

for email in emails:

    part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "ne-pas-repondre@themoocagency.com"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = email
    msg['Subject'] = "Rapport de temps passé - MOOC icope"
    attachment = _files_values
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
    msg.attach(part)
    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, email, text)
    server.quit()
    log.info('Email sent to '+str(email))


log.info('------------> Finish calculate grades and write xlsx report')


users_html = "".join(f"<li>{user}</li>" for user in buggued_tt)

html_2 = f"""<html> <head></head><body><p>Bonjour,<br/><br/> Voici la liste des utilisateurs impliqués:</p><ul>{users_html}</ul><br/><p>Bonne r&eacute;ception<br/>   L'&eacute;quipe WeUp Learning</p>  </body></html>"""

if len(buggued_tt) >=1 :

    part2 = MIMEText(html_2.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "ne-pas-repondre@themoocagency.com"
    email = "cyril.adolf@weuplearning.com"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = email
    msg['Subject'] = "Rapport TT buggué"

    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, email, text)
    server.quit()
    log.info('Email sent to '+str(email))




# refonte 2026
# sudo /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/icope/lms/utils/detailled_time_tracking_report_2026.py "melanie.zunino@weuplearning.com;dekerimel.j@chu-toulouse.fr" "course-v1:icope+e1+2026;course-v1:icope+gda+2026;course-v1:icope+ea+2026;course-v1:icope+soins_bucco_dentaires+2024"
