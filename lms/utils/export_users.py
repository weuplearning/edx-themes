# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()

from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from student.models import *
from lms.djangoapps.wul_apps.models import WulCourseEnrollment


from openpyxl import Workbook
import json
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import logging
log = logging.getLogger(__name__)


# sys.setdefaultencoding('utf8')

timestr = time.strftime("%Y_%m_%d")
timesfr = time.strftime("%d.%m.%Y")
timesfr = str(timesfr)


_title = [
    "email",
    "Nom",
    "prenom",
    "Pays",
    "Genre",
    "Année de naissance",
    "Code Postal",
    "Adresse",
    "date d'inscription",
    "dernière connexion",
    "LaPatisserie - MOOCPatisserieAFPA_S1", # colonne K
    "LaPatisserie2 - MOOCPatisserieAFPA_S2",
    "MOOC_FLE_AFPA - FLE",
    "Mets et Vins - Saison 3",
    "Les101techniquesdebase - MOOCCUISINEAFPA",
    "Les101techniquesdebase - MOOCCUISINEAFPA_S2",
    "Les101techniquesdebase - MOOCCUISINEAFPA_S3",
    "Les101techniquesdebase - Replay 2019",
    "Occitanie",
    "FLI", 
    "Patisserie 2020", # colonne U
    "Mets et vins 2020",
    "FLI 2020",
    "Cuisine 2020",
    "Mixite",
    "CPF",
    "Handicap", # colonne AA
    "TRE",
    "MATU",
    "MOOC Love Food",
    "Mooc Handicap Afpa 2022", # colonne AE
    "Au coeur de la cuisine Corse",
    "Découverte du recyclage",
    "SIAE : Du projet à la demande de financement",
    "Travailler dans l'Economie Sociale et Solidaire",
    "Temps passé"
]

_id = [
    "course-v1:afpa+LaPatisserie+MOOCPatisserieAFPA_S1", # colonne K
    "course-v1:afpa+LaPatisserie2+MOOCPatisserieAFPA_S2",
    "course-v1:afpa+MOOC_FLE_AFPA+FLE",
    "course-v1:afpa+Metsetvins+MOOCmetsetvinsAFPA_S3",
    "course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA",
    "course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA_S2",
    "course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA_S3",
    "course-v1:afpa+Les101techniquesreplay+2019",
    "course-v1:afpa+occitanie+2019_S1",
    "course-v1:afpa+MOOC_FLI+FLI_2019",
    "course-v1:afpa+La_Patisserie_Replay_2020+2020", # colonne U
    "course-v1:afpa+Mets_et_vins_replay_2020+2020",
    "course-v1:afpa+FLI+2023",
    "course-v1:afpa+replay_2020+2020",
    "course-v1:afpa+mixite+mixite_2020",
    "course-v1:afpa+CPF+CPF_2020",
    "course-v1:afpa+inclusion_sociale+2020", # colonne AA
    "course-v1:afpa+TRE_2020+2020",
    "course-v1:afpa+MATU+2020",
    "course-v1:afpa+love_food+2020",
    "course-v1:afpa+inclusion_sociale+2023", # colonne AE
    "course-v1:afpa+corse+2024",
    "course-v1:afpa+recyclage+2024",  # New 20/01/2025
    "course-v1:afpa+dreets+2024", # New 20/01/2025
    "course-v1:afpa+ess+2024" # New 20/01/2025
]

# blacklist
prunable_courses_indexes = [
    11,#"course-v1:afpa+LaPatisserie+MOOCPatisserieAFPA_S1", # colonne K
    12,#"course-v1:afpa+LaPatisserie2+MOOCPatisserieAFPA_S2", L
    #13#"course-v1:afpa+MOOC_FLE_AFPA+FLE", M
    14,#"course-v1:afpa+Metsetvins+MOOCmetsetvinsAFPA_S3", N
    15,#"course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA", O
    16,#"course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA_S2", P
    17,#"course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA_S3", Q
    18#"course-v1:afpa+Les101techniquesreplay+2019", R
    #19#"course-v1:afpa+occitanie+2019_S1",S
    #20#"course-v1:afpa+MOOC_FLI+FLI_2019",T
    #21#"course-v1:afpa+La_Patisserie_Replay_2020+2020", # colonne U
    #22#"course-v1:afpa+Mets_et_vins_replay_2020+2020",
    #23#"course-v1:afpa+FLI+2023",
    #24#"course-v1:afpa+replay_2020+2020",
    #25#"course-v1:afpa+mixite+mixite_2020",
    #26#"course-v1:afpa+CPF+CPF_2020",
    #27#"course-v1:afpa+inclusion_sociale+2020", # colonne AA
    #28#"course-v1:afpa+TRE_2020+2020",
    #29#"course-v1:afpa+MATU+2020",
    #30#"course-v1:afpa+love_food+2020",
    #31#"course-v1:afpa+inclusion_sociale+2023", # colonne AE
    #32#"course-v1:afpa+corse+2024",
    #33#"course-v1:afpa+recyclage+2024",  # New 20/01/2025
    #34#"course-v1:afpa+dreets+2024", # New 20/01/2025
    #35#"course-v1:afpa+ess+2024" # New 20/01/2025
]



def get_time_tracking(enrollment):

    try:
        wul_enrollment,is_exist=WulCourseEnrollment.objects.get_or_create(course_enrollment_edx=enrollment)
        global_time=wul_enrollment.global_time_tracking
    except:
        global_time = 0
    return global_time


def get_course_enrollment(course_id, user):

    try:
        enrollment = CourseEnrollment.objects.get(course_id=course_id, user=user)
    except:
        return 0
    return get_time_tracking(enrollment)


def prune_old_courses(sheet):

    courses_amount = len(_id)
    col_index = 11 # first column with usable data ("oui"/"non")
    maxcol = col_index + (courses_amount-1)

    # any col without "oui"
    eligible_cols = []
    # eligible and contained inside prunable_courses_indexes
    to_delete_cols = []

    cols = sheet.iter_cols(min_col=11,max_col=maxcol,min_row=2)
    
    for col in cols:
        
        row_index=2 # right after header
        col_eligible = True # flag
        for row in col:
            if(row.value=='oui'):
                col_eligible = False
                continue
            row_index+=1
        
        if((col_eligible == True) and (col_index in prunable_courses_indexes)):
            to_delete_cols.append(col_index)
        
        col_index+=1


    # delete cols, array is reversed to prevent unintentional index offset
    print("Pruning these columns :")
    print(to_delete_cols)
    for col_index in reversed(to_delete_cols):
        sheet.delete_cols(col_index)





#PREPARE LE XLS

filename = '/edx/var/edxapp/media/microsites/afpa/reports_{}_export_enroll_afpa.xlsx'.format(timestr)
wb = Workbook()
sheet = wb.active
sheet.title= 'Enroll'
for i, header in enumerate(_title):
   sheet.cell(1, i+1, header)

#BIG REQ SQL
query = '''
SELECT 
  a.id, 
  a.username, 
  a.first_name, 
  a.last_name, 
  a.email, 
  b.name, 
  b.custom_field, 
  c.user_id, 
  group_concat(c.course_id) AS course_id 
FROM 
  auth_user a, 
  auth_userprofile b, 
  student_courseenrollment c 
WHERE 
  a.id = c.user_id 
  AND a.id = b.user_id 
  AND c.course_id in (
'''

i=0
while i < len(_id) - 1:
   query = query + '"' + str(_id[i]) + '",'
   i = i + 1

query = query + '"' + str(_id[len(_id) - 1]) + '"'

query = query + ') GROUP BY a.id;'



users = User.objects.raw(query)

i = 1
print("foreach user")
for user in users:

    global_time = 0


    try:
        registration_date = user.date_joined.strftime('%d %b %y')
    except:
        registration_date = "n/a"

    try:
        last_login = user.last_login.strftime('%d %b %y')
    except:
        last_login = "n/a"


    _email = user.email

    if ( _email.find('@weuple') != -1 ) :
        continue

    try:
        _custom = json.loads(user.custom_field)
    except:
        _custom = {}


    if _custom.get('last_name') is not None:
        _last_name = _custom.get('last_name')
    else:
        try :
            _last_name = user.name.split(' ')[1]
        except :
            _last_name = user.last_name


    if _custom.get('first_name') is not None:
        _first_name = _custom.get('first_name')
    else:
        try : 
            _first_name = user.name.split(' ')[0]
        except :
            _first_name = user.first_name


    userId = User.objects.get(id=user.id)
    user_profile = UserProfile.objects.get(user_id=user.id)


    try:
        log.info(userId)
        log.info(user_profile)

        _custom = json.loads(user_profile.custom_field)

        log.info('456')
        log.info(_custom)
        if _custom.get('gender') :
            if _custom['gender'] == 'm' or _custom['gender'] == 'homme' or _custom['gender'] == 'h' :
                _custom['gender'] = 'Homme' 

        # if _custom['countr']y[0].isupper() :
        # if _custom.get('country') :
        #     _custom['country'] = _custom['country'].capitalize()
        #     if _custom['country'] == 'Fr' :
        #         _custom['country'] = 'France'


        log.info('789')
        log.info(_custom)


        user_profile.custom_field = json.dumps(_custom)
        userId.save()
        user_profile.save()

    except:
        print('error')
        print(user.email)
        break



    values = [
        _last_name,
        _first_name,
        _custom.get('country'),
        _custom.get('gender'),
        _custom.get('year_of_birth'),
        _custom.get('cp'),
        _custom.get('mailing_adress'),
        registration_date,
        last_login
    ]

    courses = user.course_id


    # ecriture colonne de chaque ligne
    sheet.cell(i+1, 1, _email)
    j = 1
    #values from custom form
    for val in values:
        if val:
            sheet.cell(i+1, j+1, str(val))
        else:
            sheet.cell(i+1, j+1, 'n/a')
        j = j + 1
    #is enroll
    q = {}
    course_id = user.course_id.split(',')

    for n in _id:
        for c in course_id:
            if str(n) == str(c):
                q[n] = True


    for n in _id:
        try:
            q[n]
            sheet.cell(i+1, j+1, 'oui')
        except:
            sheet.cell(i+1, j+1, 'non')
        j = j + 1
    i = i + 1

    for course in course_id:
        global_time += get_course_enrollment(course, user)

    sheet.cell(i, j+1, global_time//60)


print("pruning unused courses")
prune_old_courses(sheet)

wb.close()
output = BytesIO()
wb.save(output)
_files_values = output.getvalue()


html = u"<html><head></head><body><p>Bonjour,<br/><br/>Voici la liste des inscrits Afpa.<br/><br/>Bonne reception<br>L'équipe WeUp Learning<br></p></body></html>"
part2 = MIMEText(html, 'html')
TO_EMAILS = sys.argv[1].split(";")

for i in range(len(TO_EMAILS)):
   fromaddr = "no-reply@themoocagency.com"
   toaddr = str(TO_EMAILS[i])
   msg = MIMEMultipart()
   msg['From'] = fromaddr
   msg['To'] = toaddr
   msg['Subject'] = "Inscriptions MOOC AFPA"
   attachment = _files_values
   part = MIMEBase('application', 'octet-stream')
   part.set_payload((attachment))
   encoders.encode_base64(part)
   part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
   msg.attach(part)
   server = smtplib.SMTP('mail3.themoocagency.com', 25)
   server.starttls()
   server.login('contact', 'waSwv6Eqer89')
   msg.attach(part2)
   text = msg.as_string()
   server.sendmail(fromaddr, toaddr, text)
   server.quit()
   print('mail sent to '+str(TO_EMAILS[i]))



# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/afpa/lms/utils/export_users.py "loic.tournedouet@gmail.com;loic.tournedouet@afpa.fr;melanie.zunino@weuplearning.com"

# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/afpa/lms/utils/export_users.py "cyril.adolf@weuplearning.com"



