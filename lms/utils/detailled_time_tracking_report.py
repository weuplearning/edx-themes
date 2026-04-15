# -*- coding: utf-8 -*-
#!/usr/bin/env python
import importlib
import sys
importlib.reload(sys)
import os
from io import BytesIO
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.production")
os.environ.setdefault("LMS_CFG", "/edx/etc/lms.yml")
os.environ.setdefault("lms.envs.production,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()


#############################################################
#         ^ SETUP ENVIRONNEMENT VARIABLE FOR KOA ^          #
#                START BEYOND THIS LINE                     #
#############################################################


import json
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font



from opaque_keys.edx.locator import CourseLocator
from common.djangoapps.student.models import CourseEnrollment
from lms.djangoapps.courseware.courses import get_course_by_id
from lms.djangoapps.wul_apps.best_grade.helpers import check_best_grade
from lms.djangoapps.wul_apps.models import WulCourseEnrollment

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


import logging
log = logging.getLogger()


emails = sys.argv[1].split(";")
course_ids = sys.argv[2].split(";")

regions = ["Auvergne-rhone-alpes","bourgogne-franche-comte","Bretagne","Centre-val-de-loire","Corse","Grand-est","Hauts-de-france","Ile-de-france","Normandie","Nouvelle-aquitaine","Occitanie","Pays-de-la-loire","Provence-alpes-cote-d-azur","Guadeloupe","Martinique","Guyane","La-reunion","Mayotte","Autre"]


all_users_data = {}
headers = ['Email', 'Nom complet', 'Adresse', 'Code postal', 'Ville',  'Région', 'Département', 'Parcours', 'Profession', 'Profession si autre', 'Jour de connexion et temps', 'Temps passé par module', 'Progression']

for course_id in course_ids:
    course_key = CourseLocator.from_string(course_id)
    course = get_course_by_id(course_key)
    course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)
    course_name = course.display_name_with_default
    course_data = {}

    for i in range(len(course_enrollments)):
        print(f'{i}/{str(len(course_enrollments))}')
        user = course_enrollments[i].user




        # Replace str to get target email faster
#        if str(user.email) != "celinecoquett@gmail.com"  :
#            continue






        user_data = {}
        enrollment = course_enrollments[i]

        if str(user.email).find('@yopmail') != -1 or str(user.email).find('@weuplearning') != -1 or str(user.email).find('@themoocagency') != -1 :
            continue





        # TimeTracking
        try:
            wul_course_enrollment = WulCourseEnrollment.objects.get(course_enrollment_edx__user=user, course_enrollment_edx__course_id=course_key)
            daily_time_tracking = json.loads(wul_course_enrollment.daily_time_tracking)
            detailled_time_tracking = json.loads(wul_course_enrollment.detailed_time_tracking)
        except:
            # Pas besoin de traiter l'utilisateur 
            continue

        if daily_time_tracking == {} :
            continue



        # Profile info
        not_found_str = 'n.a.'
        user_data["email"] = user.email
        user_data["name"] = user.profile.name
        user_data["adress"] = json.loads(user.profile.custom_field).get('adress',not_found_str)
        user_data["post_code"] = json.loads(user.profile.custom_field).get('post_code',not_found_str)
        user_data["city"] = json.loads(user.profile.custom_field).get('city',not_found_str)
        user_data["region"] = json.loads(user.profile.custom_field).get('region',not_found_str)
        user_data["department"] = json.loads(user.profile.custom_field).get('department',not_found_str)
        user_data["parcours"] = json.loads(user.profile.custom_field).get('parcours',not_found_str)    
        user_data["profession"] = json.loads(user.profile.custom_field).get('profession',not_found_str)
        user_data["profession_autre"] = json.loads(user.profile.custom_field).get('profession_autre',not_found_str)
        # user_data["icope_emailing"] = 'Vrai' if json.loads(user.profile.custom_field).get('icope_emailing', 'false') == 'true' else 'Faux'



        # Grade
        user_grade = {}
        gradesTest = check_best_grade(user, course, force_best_grade=True)
        userPersentGrade = gradesTest.summary['percent']

        try:
            user_grade['global'] = round(userPersentGrade*100,2) 
        except:
            user_grade['global'] = 0



        data = { "profil": user_data, "grades": user_grade, "tt_detailled" : detailled_time_tracking, "tt_daily" :  daily_time_tracking}
        course_data[str(user.id)]= data

    all_users_data[course_id]= course_data





correspondance_section_tt = {
    "e73e91fe60fc450789f0a5faf244143a" : "Introduction",
    "1939587bf96b484186bb524099cad8f5" : "Etape 1 : Dépistage",
    "d53973250d4f463d99e10bcaf7f0a95c" : "Etape 2 : Gestion des alertes",
    "4a22fe0f40e94dfbaf6edd6d1250261c" : "Etape 2 : Evaluation approfondie",
    "ae38e833526c4779909f40ca4e8d24eb" : "Cas clinique",
    "8ee6bd2d4c2a488fab130ad36fe67156" : "Présentation de l'outil de coordination régional",
    "a6bdc40c399b4c66ad869d216b0dc7ce" : "Conclusion"
}
if 'course-v1:audition-icope+med+01' in course_ids : 
    correspondance_section_tt = {
    # audition-icope
        "f7f3e683f4d74b3f8baa560e5c71f1e8" : "Physiopathologie de l'oreille",
        "a5981a4c949347c98fe169cb88f59134" : "Du dépistage de la surdité à la décision",
        "1c97afb70dce4ebaa1c5848df2daae98" : "Environnement des mesures auditives",
        "2a8fa472c3974e64ad940aff67abcf58" : "Acoumétrie",
        "01f282202fd547c696adaff0b07d6d08" : "Audiométrie tonale avec assourdissement",
        "f20382e487fd49ff9cf3d1ade5ddb132" : "Audiométrie vocale dans le silence et corrélation tonale/vocale",
        "96f431656590437a8e6981d09370e4d3" : "Audiométrie vocale dans le Bruit",
        "f4e42a79c12f47ba83e287d226474d42" : "Formation à l'otoscopie et à la gestion des difficultés",
        "d13557ba965f4ad398831929ba361960" : "Orientation thérapeutique",
        "e832ba93bd90442c8f280e0f4bbd6240" : "Orientation vers la gériatrie",
        "0c61635589cf4f439babe61c7b62de45" : "Serious Game",
        "8d5a1a2021cb47f89850178aa0849f35" : "Evaluation finale"
    }


# WRITE EXCEL AND SEND MAILS
timestr = time.strftime("%Y_%m_%d")


# I/ Un rapport par région : 
#for region in regions :
#
#    wb = Workbook()
#    sheet = wb.active
#    sheet.title= 'Rapport de notes'
#    filename = '/home/edxtma/csv/Icope_grade_report_{}_{}.xlsx'.format(timestr, region)
#    j=1
#    for k, course_id in all_users_data.items():
#        for key, user in course_id.items():
#            if user['profil']['region'] == region : 
#                for i, header in enumerate(headers):
#                    sheet.cell(j, i+1, header)
#                    sheet.cell(j, i+1).fill = PatternFill("solid", fgColor="6B9AAF")
#                    sheet.cell(j, i+1).font = Font(b=False, color="FFFFFF")
#                i=0
#                j+=1
#                for key, value in  user['profil'].items() : 
#                    sheet.cell(j, i+1, value)
#                   i+=1
#                percent_global = str(user['grades']['global']) + '%'
#                sheet.cell(j, i+3, percent_global)
#                for id, section_name in correspondance_section_tt.items() : 
#                    for hash, seconds in user['tt_detailled'].items() :
#                        if hash == id :
#                            sheet.cell(j, i+2, str(section_name) + " - " + str(round(seconds/60))+" min")
#                            break
#                        else:
#                            sheet.cell(j, i+2, str(section_name) + " - 0 min")
#                    j+=1
#                j -= 7
#                for day, seconds in user['tt_daily'].items() : 
#                    sheet.cell(j, i+1, str(day) + " : " + str(round(seconds/60))+" min")
#                    j+=1
#                if len(user['tt_daily'].items()) >= 7 :
#                    j+= 1
#                else : 
#                    j+= 8 - len(user['tt_daily'].items())  
#    if j <=2 : 
#        continue

#    output = BytesIO()
#    wb.save(output)
#    _files_values = output.getvalue()

#    html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de temps passé pour la région : "+ region +"<br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

#    for email in emails:
#        part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
#        fromaddr = "ne-pas-repondre@themoocagency.com"
#        msg = MIMEMultipart()
#        msg['From'] = fromaddr
#        msg['To'] = email
#        msg['Subject'] = "icope_grade_report"
#        attachment = _files_values
#        part = MIMEBase('application', 'octet-stream')
#        part.set_payload(attachment)
#        encoders.encode_base64(part)
#        part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
#        msg.attach(part)
#        server = smtplib.SMTP('mail3.themoocagency.com', 25)
#        server.starttls()
#        server.login('contact', 'waSwv6Eqer89')
#        msg.attach(part2)
#        text = msg.as_string()
#        server.sendmail(fromaddr, email, text)
#        server.quit()
#        log.info('Email sent to '+str(email))





# II/ Un rapport toutes régions confondues :
wb = Workbook()
sheet = wb.active
sheet.title= 'Rapport de notes'
filename = '/home/edxtma/csv/Icope_grade_report_{}.xlsx'.format(timestr)


for i, header in enumerate(headers):
    sheet.cell(1, i+1, header)
    sheet.cell(1, i+1).fill = PatternFill("solid", fgColor="59C4C6")
    sheet.cell(1, i+1).font = Font(b=False, color="FFFFFF")

j=1
for k, course_id in all_users_data.items():

    for key, user in course_id.items():


        for i, header in enumerate(headers):
            sheet.cell(j, i+1, header)
            sheet.cell(j, i+1).fill = PatternFill("solid", fgColor="6B9AAF")
            sheet.cell(j, i+1).font = Font(b=False, color="FFFFFF")

        i=0
        j+=1

        for key, value in  user['profil'].items() : 
            sheet.cell(j, i+1, value)
            i+=1


        percent_global = str(user['grades']['global']) + '%'
        sheet.cell(j, i+3, percent_global)

        for id, section_name in correspondance_section_tt.items() : 
            for hash, seconds in user['tt_detailled'].items() : 

                if hash == id :
                    sheet.cell(j, i+2, str(section_name) + " - " + str(round(seconds/60))+" min")
                    break
                else:
                    sheet.cell(j, i+2, str(section_name) + " - 0 min")

            j+=1


        j -= 7

        for day, seconds in user['tt_daily'].items() : 

            sheet.cell(j, i+1, str(day) + " : " + str(round(seconds/60))+" min")
            j+=1


        if len(user['tt_daily'].items()) >= 7 :
            j+= 1
        else : 
            j+= 8 - len(user['tt_daily'].items())  


output = BytesIO()
wb.save(output)
_files_values = output.getvalue()

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de temps passé regroupant toutes les régions <br/><br/>Bonne r&eacute;ception<br/>L'&eacute;quipe WeUp Learning</p></body></html>"

for email in emails:

    part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')
    fromaddr = "ne-pas-repondre@themoocagency.com"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = email
    msg['Subject'] = "icope_grade_report"
    attachment = _files_values
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
    msg.attach(part)
    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    server.sendmail(fromaddr, email, text)
    server.quit()
    log.info('Email sent to '+str(email))


log.info('------------> Finish calculate grades and write xlsx report')




# Qualif
# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/icope/lms/utils/detailled_time_tracking_report.py 'cyril.adolf@weuplearning.com' course-v1:icope+Occitanie+2022


# Prod
# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/icope/lms/utils/detailled_time_tracking_report.py 'cyril.adolf@weuplearning.com' "course-v1:icope+Occitanie+2022;course-v1:icope+Centre_Val_de_Loire+2022;course-v1:icope+Corse+2022;course-v1:icope+La_Reunion+2022;course-v1:icope+Auvergne_Rhone_Alpes+2022;course-v1:icope+Pays_de_la_Loire+2022;course-v1:icope+PACA+2022;course-v1:icope+Grand_Est+2022;course-v1:icope+Nouvelle_Aquitaine+2022"


# Audition-icope :
# /edx/app/edxapp/venvs/edxapp/bin/python /edx/app/edxapp/edx-themes/icope/lms/utils/detailled_time_tracking_report.py 'cyril.adolf@weuplearning.com' course-v1:audition-icope+med+01


